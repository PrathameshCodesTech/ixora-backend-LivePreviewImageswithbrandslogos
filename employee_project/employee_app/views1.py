import os
import subprocess
from rest_framework.decorators import api_view,parser_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from rest_framework import viewsets
from rest_framework.views import APIView
from django.core.files import File
from django.core.files.storage import default_storage
from datetime import timedelta
from django.db import IntegrityError
from .models import Employee ,DoctorVideo, Doctor,EmployeeLoginHistory, VideoTemplates,DoctorOutputVideo,ImageContent
from .serializers import EmployeeLoginSerializer,EmployeeSerializer,DoctorSerializer,DoctorVideoSerializer, VideoTemplatesSerializer,DoctorOutputVideoSerializer,ImageContentSerializer,ImageTemplateSerializer
import pandas as pd # type: ignore
from django.conf import settings
from django.core.files import File
from django.db.models import Q, Count   
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from rest_framework_simplejwt.views import TokenRefreshView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.response import Response
from rest_framework import status
import requests
import uuid
from distutils.util import strtobool
from rest_framework.permissions import IsAuthenticated
from .serializers import DoctorVideoSerializer
from datetime import datetime, date
import random
import string
import openpyxl
from django.utils import timezone
from django.http import HttpResponse
import logging
from rest_framework.pagination import PageNumberPagination
import uuid
from PIL import Image, ImageDraw, ImageFont  
from django.http import Http404





# Set up logger
logger = logging.getLogger(__name__)

def parse_css_shadow(shadow_str):
    """Parse CSS text-shadow: '2px 2px 4px rgba(0,0,0,0.7)'"""
    if shadow_str == 'none' or not shadow_str:
        return None
    
    try:
        # Handle rgba colors
        if 'rgba' in shadow_str:
            rgba_start = shadow_str.find('rgba(')
            rgba_end = shadow_str.find(')', rgba_start)
            rgba_str = shadow_str[rgba_start+5:rgba_end]
            rgba_values = [float(x.strip()) for x in rgba_str.split(',')]
            shadow_color = (int(rgba_values[0]), int(rgba_values[1]), int(rgba_values[2]))
            
            offset_part = shadow_str[:rgba_start].strip()
            offsets = offset_part.replace('px', '').split()
            
            if len(offsets) >= 2:
                return {
                    'offset_x': int(float(offsets[0])), 
                    'offset_y': int(float(offsets[1])), 
                    'color': shadow_color
                }
        elif 'px' in shadow_str:
            parts = shadow_str.replace('px', '').split()
            if len(parts) >= 2:
                return {
                    'offset_x': int(float(parts[0])), 
                    'offset_y': int(float(parts[1])), 
                    'color': (128, 128, 128)
                }
    except:
        pass
    
    return {'offset_x': 2, 'offset_y': 2, 'color': (128, 128, 128)}

class Pagination_class(PageNumberPagination):
    page_size = 10  # You can change this default page size
    page_size_query_param = 'page_size'
    max_page_size = 100

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer

def get_tokens_for_employee(employee):
    refresh = RefreshToken.for_user(employee)
    
    # Set expiration time for access token
    refresh.access_token.set_exp(lifetime=timedelta(hours=1))  # Expiry in 1 hour

    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
        'access_token_exp': refresh.access_token.payload['exp'],  # Expiry time of access token
    }

@api_view(['POST'])
def employee_login_api(request):
    serializer = EmployeeLoginSerializer(data=request.data)
    if serializer.is_valid():
        employee_id = serializer.validated_data['employee_id']
        try:
            employee = Employee.objects.get(employee_id=employee_id)

            if not employee.status:
                return Response({
                    'status': 'error',
                    'message': 'Your account is inactive. Please contact the admin department.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            employee.login_date = timezone.now()
            employee.has_logged_in = True
            employee.save(update_fields=['login_date', 'has_logged_in']) 
            
            
            tokens = get_tokens_for_employee(employee)

            EmployeeLoginHistory.objects.create(
                employee=employee,
                employee_identifier=employee.employee_id,
                name=f"{employee.first_name} {employee.last_name}",
                email=employee.email,
                phone=employee.phone,
                department=employee.department,
                user_type=employee.user_type
            )


            return Response({
                'status': 'success',
                'message': 'Login successful',
                'tokens': tokens,
                'employee': {
                    'id':employee.id,
                    'employee_id': employee.employee_id,
                    'name': f"{employee.first_name} {employee.last_name}",
                    'email': employee.email,
                    'department': employee.department,
                    'user_type': employee.user_type
                }
            }, status=status.HTTP_200_OK)

        except Employee.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Invalid employee ID'
            }, status=status.HTTP_401_UNAUTHORIZED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])  # Important for file uploads
def add_doctor(request):
    serializer = DoctorSerializer(data=request.data)
    # print(serializer,"serializer")
    if serializer.is_valid():
        serializer.save()
        return Response({'status': 'success', 'data': serializer.data}, status=status.HTTP_201_CREATED)
    return Response({'status': 'error', 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


###### Video creation #######

class VideoGenViewSet(viewsets.ModelViewSet):
    queryset = DoctorVideo.objects.all()
    serializer_class = DoctorVideoSerializer

    # def perform_create(self, serializer):
    #     request = self.request
    #     print(f"🔍 REQUEST DATA: {request.data}")
    #     print(f"🔍 REQUEST FILES: {request.FILES}")
    #     doctor = serializer.save()
    #     print(f"🔍 DOCTOR SAVED SUCCESSFULLY: {doctor.id}")
    #     print(f"🔍 DOCTOR CREATED: ID={doctor.id}, Name={doctor.name}")
    #     print(f"🔍 DOCTOR HAS IMAGE: {bool(doctor.image)}")
        
    #     print(f"🔍 DJANGO VIDEO: Doctor created with ID: {doctor.id}")
    #     print(f"🔍 DJANGO VIDEO: Doctor name: {doctor.name}")
    #     print(f"🔍 DJANGO VIDEO: Has image: {bool(doctor.image)}")
        
    #     # Skip if doctor has no image
    #     if not doctor.image:
    #         print(f"🔍 DJANGO VIDEO: Skipping video generation - no image")
    #         return

    #     # Use template_id from request if given
    #     template_id = request.data.get("template_id")
    #     template = None

    #     if template_id:
    #         template = VideoTemplates.objects.filter(id=template_id).first()
    #     else:
    #         # Fallback to default template (status=True)
    #         template = VideoTemplates.objects.filter(status=True).first()

    #     # If no valid template, skip video creation
    #     if not template:
    #         return

    #     # Prepare output path
    #     output_filename = f"{doctor.id}_{template.id}_output.mp4"
    #     output_dir = os.path.join(settings.MEDIA_ROOT, "output", str(doctor.employee.id), str(doctor.id))
    #     os.makedirs(output_dir, exist_ok=True)
    #     output_path = os.path.join(output_dir, output_filename)

    #     try:
    #         # Call the same logic from GenerateDoctorOutputVideoView
    #         self.generate_custom_video(
    #             main_video_path=template.template_video.path,
    #             image_path=doctor.image.path,
    #             name=doctor.name,
    #             clinic=doctor.clinic,
    #             city=doctor.city,
    #             specialization_key=doctor.specialization_key,
    #             state=doctor.state,
    #             output_path=output_path
    #         )

    #         # Save generated video
    #         relative_path = os.path.relpath(output_path, settings.MEDIA_ROOT)

    #         DoctorOutputVideo.objects.create(
    #             doctor=doctor,
    #             template=template,
    #             video_file=relative_path
    #         )
    #     except Exception as e:
    #         print(f"Error generating video after doctor creation: {e}")

    def perform_create(self, serializer):
        print("🔍 ===== PERFORM_CREATE CALLED =====")
        request = self.request
        print(f"🔍 REQUEST DATA: {dict(request.data)}")
        doctor = serializer.save()
        
        print(f"🔍 DOCTOR CREATED: ID={doctor.id}, Name={doctor.name}")
        print(f"🔍 - Has Image: {bool(doctor.image)}")
        print(f"🔍 - Image Path: {doctor.image.path if doctor.image else 'None'}")
        print(f"🔍 DOCTOR HAS IMAGE: {bool(doctor.image)}")

        # Skip if doctor has no image
        if not doctor.image:
            print("🔍 NO IMAGE, SKIPPING VIDEO")
            return
        else:
            print(f"🔍 ✅ IMAGE FOUND: {doctor.image.path}")


        # Use template_id from request if given
        template_id = request.data.get("template_id")
        print(f"🔍 TEMPLATE ID FROM REQUEST: {template_id}")
        template = None

        if template_id:
            template = VideoTemplates.objects.filter(id=template_id).first()
            print(f"🔍 USING TEMPLATE: {template.name if template else 'NOT FOUND'}")
        else:
            # Fallback to default template (status=True)
            template = VideoTemplates.objects.filter(status=True).first()
            print(f"🔍 USING DEFAULT TEMPLATE: {template.name if template else 'NOT FOUND'}")

        # If no valid template, skip video creation
        if not template:
            print("🔍 NO TEMPLATE FOUND, SKIPPING VIDEO")
            return

        # Prepare output path
        output_filename = f"{doctor.id}_{template.id}_output.mp4"
        output_dir = os.path.join(settings.MEDIA_ROOT, "output", str(doctor.employee.id), str(doctor.id))
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, output_filename)
        
        print(f"🔍 OUTPUT PATH: {output_path}")

        try:
            # Call video generation
            self.generate_custom_video(
                main_video_path=template.template_video.path,
                image_path=doctor.image.path,
                name=doctor.name,
                clinic=doctor.clinic,
                city=doctor.city,
                specialization_key=doctor.specialization_key,
                state=doctor.state,
                output_path=output_path
            )

            # Save generated video
            relative_path = os.path.relpath(output_path, settings.MEDIA_ROOT)
            
            DoctorOutputVideo.objects.create(
                doctor=doctor,
                template=template,
                video_file=relative_path
            )
            
            print(f"🔍 VIDEO GENERATION SUCCESS: {relative_path}")
            
        except Exception as e:
            print(f"🔍 ERROR GENERATING VIDEO: {e}")
            import traceback
            traceback.print_exc()
    

    def generate_custom_video(self, main_video_path, image_path, name, clinic, city, specialization_key, state, output_path):
        temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
        os.makedirs(temp_dir, exist_ok=True)

        resolution = "415x410"
        fps = 30
        fade_duration = 3
        slots = [(2, 6), (65, 70)]
        
        temp_videos = []
        for i, (start, end) in enumerate(slots):
            duration = end - start
            total_frames = duration * fps

            zoom_effect = f"zoompan=z='1+0.00003*in':x='(iw/2)-(iw/zoom/2)':y='(ih/2)-(ih/zoom/2)':d={total_frames}:s={resolution}:fps={fps}"
            fade_effect = f"format=rgba,fade=t=in:st=0:d={fade_duration}:alpha=1,fade=t=out:st={duration-fade_duration}:d={fade_duration}:alpha=1"
            vf = f"scale={resolution},{zoom_effect},{fade_effect}"

            temp_video = os.path.join(temp_dir, f"temp_image_vid_{i}.mp4")

            subprocess.run([
                r'C:\ffmpeg\bin\ffmpeg.exe', "-loop", "1", "-i", image_path,
                "-vf", vf, "-t", str(duration), "-y", temp_video
            ], check=True)

            temp_videos.append((temp_video, start, end))

        # === Drawtext settings ===
        text_lines = [name, specialization_key, clinic, city, state]
        font = "RobotoSlab-Medium.ttf"
        
        text_filters = []
        for start, end in slots:
            alpha_expr = f"if(lt(t\\,{start}+3),(t-{start})/3,if(lt(t\\,{end}-3),1,({end}-t)/3))"
            for j, text in enumerate(text_lines):
                offset = 132
                y_pos = f"(main_h - ({len(text_lines)}*60) + {j}*60 - {offset})"
                x_pos = "(main_w/2)-300"
            
                drawtext = (
                    f"drawtext=text='{text}':fontfile='{font}':fontcolor=black:fontsize=40:"
                    f"x={x_pos}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
                )
                text_filters.append(drawtext)

        overlay_x1 = f"(main_w-overlay_w)/2-350"
        overlay_y1 = f"(main_h-overlay_h)/2+70"

        filter_complex = (
            f"[0:v][1:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{slots[0][0]},{slots[0][1]})'[v1];"
            f"[v1][2:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{slots[1][0]},{slots[1][1]})'[v2];"
            f"[v2]{','.join(text_filters)}[v]"
        )

        cmd = [
            r'C:\ffmpeg\bin\ffmpeg.exe', "-i", main_video_path,
            "-i", temp_videos[0][0],
            "-i", temp_videos[1][0],
            "-filter_complex", filter_complex,
            "-map", "[v]",
            "-map", "0:a?",
            "-c:v", "libx264", "-c:a", "copy", "-y", output_path
        ]

        subprocess.run(cmd, check=True)

        # Clean temp files
        for temp_video, _, _ in temp_videos:
            os.remove(temp_video)

    # def perform_create(self, serializer):
    #     instance = serializer.save()

    #     print(instance,"instance")

    #     start = 10
    #     end = 15

    #     image_path = instance.image.path
    #     output_dir = os.path.join(settings.MEDIA_ROOT, "output")

    #     os.makedirs(output_dir, exist_ok=True)
    #     output_path = os.path.join(output_dir, f"{instance.id}_output.mp4")

    #     # main_video_path = "/home/akshay/Vibe_doctors/employee_project/employee_app/media/docter_24_5.mp4"
    #     # main_video_path = "/home/ubuntu/Vibe_doctors/employee_project/employee_app/media/docter_24_5.mp4"
    #     main_video_path = os.path.join(settings.MEDIA_ROOT,"s1233.mp4")

    #     try:
    #         self.generate_custom_video(
    #             main_video_path,
    #             image_path,
    #             instance.name,
    #             instance.clinic,
    #             instance.city,
    #             instance.specialization,
    #             instance.state,

                
    #             output_path
    #         )



    #         # Attach output video file

    #         BASE_URL = "https://api.videomaker.digielvestech.in/"



    #         with open(output_path, 'rb') as f:
    #             instance.output_video.save(f"{instance.id}_output.mp4", File(f), save=True)
            
    #         instance.refresh_from_db()

    #         # current_site = Site.objects.get_current()
    #         # instance.output_video_url = f"http://{current_site.domain}{instance.output_video.url}"
    #         # BASE_URL = "http://localhost:8000"
            
    #         # BASE_URL = "http://13.126.205.205:8002"

    #         instance.output_video_url = f"{BASE_URL}{instance.output_video.url}"
    #         instance.save()


    #     except Exception as e:
    #         print(f"Error generating video: {e}")

    # def generate_custom_video(self, main_video_path, image_path, name, clinic, city,specialization, state, output_path):

    #     temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
    #     os.makedirs(temp_dir, exist_ok=True)
    
    #     resolution = "300x300"
    #     fps = 30
    #     fade_duration = 3
    
    #     # slots = [(10, 15), (43, 50)] # Time intervals
    #     slots = [(10, 15),(61,65)]
        
    #     temp_videos = []
    #     for i, (start, end) in enumerate(slots):
    #         duration = end - start
    #         total_frames = duration * fps
    
    #         zoom_effect = f"zoompan=z='1+0.00003*in':x='(iw/2)-(iw/zoom/2)':y='(ih/2)-(ih/zoom/2)':d={total_frames}:s={resolution}:fps={fps}"
    #         fade_effect = f"format=rgba,fade=t=in:st=0:d={fade_duration}:alpha=1,fade=t=out:st={duration-fade_duration}:d={fade_duration}:alpha=1"
    #         vf = f"scale={resolution},{zoom_effect},{fade_effect}"
    
    #         temp_video = os.path.join(temp_dir, f"temp_image_vid_{i}.mp4")
    
    #         subprocess.run([
    #             "ffmpeg", "-loop", "1", "-i", image_path,
    #             "-vf", vf, "-t", str(duration), "-y", temp_video
    #         ], check=True)
    
    #         temp_videos.append((temp_video, start, end))
    
    #     # === Prepare drawtext ===
    #     text_lines = [name, specialization, clinic, city, state]
    #     font = "RobotoSlab-Medium.ttf"
    #     # text_x = "(main_w/2)+180"
    #     # base_y = "(main_h/2)-150"
    #     text_x = "(main_w/2)-150"
    #     base_y = "(main_h/2)-60"
    #     line_spacing = 60
    
    #     text_filters = []
    #     for start, end in slots:
    #         alpha_expr = f"if(lt(t\\,{start}+3),(t-{start})/3,if(lt(t\\,{end}-3),1,({end}-t)/3))"
    #         for j, text in enumerate(text_lines):
    #             y_pos = f"{base_y}+{j}*{line_spacing}"
    #             drawtext = (
    #                 f"drawtext=text='{text}':fontfile='{font}':fontcolor=black:fontsize=40:"
    #                 f"x={text_x}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
    #             )
    #             text_filters.append(drawtext)
    
    #     # overlay_x = "(main_w-overlay_w)/2-150"
    #     # overlay_y = "(main_h-overlay_h)/2+30"


    #     overlay_x = "(main_w-overlay_w)/2-350"
    #     overlay_y = "(main_h-overlay_h)/2+70"
    
    #     # === Final filter_complex ===
    #     filter_complex = (
    #         f"[0:v][1:v]overlay=x={overlay_x}:y={overlay_y}:enable='between(t,{slots[0][0]},{slots[0][1]})'[v1];"
    #         f"[v1][2:v]overlay=x={overlay_x}:y={overlay_y}:enable='between(t,{slots[1][0]},{slots[1][1]})'[v2];"
    #         f"[v2]{','.join(text_filters)}[v]"
    #     )
        
    #     cmd = [
    #         "ffmpeg", "-i", main_video_path,
    #         "-i", temp_videos[0][0],
    #         "-i", temp_videos[1][0],
    #         "-filter_complex", filter_complex,
    #         "-map", "[v]",
    #         "-map", "0:a?",  # optional audio
    #         "-c:v", "libx264", "-c:a", "copy", "-y", output_path
    #     ]
        
    
    #     subprocess.run(cmd, check=True)
    
    #     # Clean temp
    #     for temp_video, _, _ in temp_videos:
    #         os.remove(temp_video)



class DoctorVideoViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = DoctorVideo.objects.all().order_by('-created_at')
    serializer_class = DoctorVideoSerializer
    pagination_class = Pagination_class



@api_view(['POST'])
@parser_classes([MultiPartParser])
def bulk_upload_employees(request):
    excel_file = request.FILES.get('file')
    if not excel_file:
        return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        df = pd.read_excel(excel_file)
        required_columns = {'first_name', 'last_name', 'email', 'phone', 'department', 'date_joined'}

        if not required_columns.issubset(df.columns):
            return Response({'error': f'Missing required columns: {required_columns - set(df.columns)}'}, status=status.HTTP_400_BAD_REQUEST)

        created, skipped, errors = 0, 0, []

        def generate_employee_id(first_name, last_name):
            base = f"EMP{first_name[0].upper()}{last_name[0].upper()}"
            while True:
                suffix = ''.join(random.choices(string.digits, k=4))
                emp_id = f"{base}{suffix}"
                if not Employee.objects.filter(employee_id=emp_id).exists():
                    return emp_id

        for index, row in df.iterrows():
            row_number = index + 2

            first_name = str(row.get('first_name', '')).strip()
            last_name = str(row.get('last_name', '')).strip()
            email = str(row.get('email', '')).strip()
            phone = str(row.get('phone', '')).strip()

            if not first_name or not last_name or not email:
                errors.append({'row': row_number, 'error': 'Missing first name, last name, email or phone'})
                skipped += 1
                continue

            employee_id = str(row.get('employee_id')).strip()

            data = {
                'employee_id': employee_id,
                'first_name': first_name,
                'last_name': last_name,
                'email': email,
                'phone': phone,
                'department': row.get('department', '') or 'Unknown',
                'date_joined': row.get('date_joined') if not pd.isna(row.get('date_joined')) else datetime.today().date()
            }

            serializer = EmployeeSerializer(data=data)
            if serializer.is_valid():
                try:
                    serializer.save()
                    created += 1
                except IntegrityError:
                    errors.append({'row': row_number, 'error': f'Duplicate email or other unique constraint'})
                    skipped += 1
            else:
                errors.append({'row': row_number, 'error': serializer.errors})
                skipped += 1

        return Response({
            'created': created,
            'skipped': skipped,
            'errors': errors
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


   
class DoctorListByEmployee(APIView):

    permission_classes = [IsAuthenticated]
      
    def get(self, request):
        # Get the employee_id from query parameters
        employee_id = request.GET.get('employee_id')

        if not employee_id:
            return Response({"detail": "employee_id is required."}, status=400)

        try:
            # Find the employee by employee_id
            employee = Employee.objects.get(employee_id=employee_id)
        except Employee.DoesNotExist:
            return Response({"detail": "Employee not found for this employee_id."}, status=404)

        # Filter doctors associated with the employee
        doctors = Doctor.objects.filter(employee=employee)

        # Serialize the data
        serializer = DoctorSerializer(doctors, many=True)
        return Response(serializer.data)
    

class DoctorVideoListView(APIView):

    def get(self, request):
        employee_id = request.GET.get('employee_id')

        if not employee_id:
            return Response({"detail": "employee_id is required."}, status=400)

        try:
            employee = Employee.objects.get(employee_id=employee_id)
        except Employee.DoesNotExist:
            return Response({"detail": "Employee not found for this employee_id."}, status=404)

        doctors = DoctorVideo.objects.filter(employee=employee).order_by('-created_at')  # sort by newest first

        # Apply pagination
        paginator = Pagination_class()
        page = paginator.paginate_queryset(doctors, request)

        serializer = DoctorVideoSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)

class DoctorVideoGeneration(APIView):

    def post(self, request):
        # Use request.GET.get('id') in production — hardcoded for testing here
        doctor_id = request.data.get('id')

        if not doctor_id:
            return Response({"detail": "doctor_id is required."}, status=400)

        try:
            # Fetch a single DoctorVideo instance
            doctor_data = DoctorVideo.objects.get(id=doctor_id)
            print(doctor_data, "doctor_data-------------")
            # Call your video generation logic
            generate_video_for_doctor(doctor_data)

            return Response({
                "detail": "Video generation successful.",
                "video_path": request.build_absolute_uri(doctor_data.output_video.url) if doctor_data.output_video else None
            })

        except DoctorVideo.DoesNotExist:
            return Response({"detail": "Doctor not found for this doctor_id."}, status=404)

        except Exception as e:
            return Response({"detail": f"Error during video generation: {str(e)}"}, status=500)


class CustomTokenRefreshView(TokenRefreshView):
    def post(self, request, *args, **kwargs):
        refresh_token = request.data.get('refresh', None)
        if not refresh_token:
            return Response({
                'status': 'error',
                'message': 'Refresh token is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Verify the refresh token and get new access token
            refresh = RefreshToken(refresh_token)

            print(refresh)

            new_refresh_token = str(refresh)

            new_access_token = str(refresh.access_token)

            return Response({
                'status': 'success',
                'message': 'Token refreshed successfully',
                'access': new_access_token,
                'refresh': new_refresh_token,
                'access_token_exp': refresh.access_token.payload['exp'],  # Expiry time of access token
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'status': 'error',
                'message': 'Invalid refresh token'
            }, status=status.HTTP_401_UNAUTHORIZED)
        



def generate_video_for_doctor(doctor):
    print(f"Generating video for doctor {doctor.name}")

    try:
        if not doctor.image:
            print(f"No image for doctor {doctor.name}, skipping video.")
            return

        # Use simple hardcoded template path first
        main_video_path = os.path.join(settings.MEDIA_ROOT, "Health111.mp4")
        image_path = doctor.image.path
        print(image_path,"image_path")
        output_dir = os.path.join(settings.MEDIA_ROOT, "output")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{doctor.id}_output.mp4")

        # Use the same method from VideoGenViewSet
        video_gen = VideoGenViewSet()
        video_gen.generate_custom_video(
            main_video_path,
            image_path,
            doctor.name,
            doctor.clinic,
            doctor.city,
            doctor.specialization,
            doctor.state,
            output_path
        )

        # Save the output video
        with open(output_path, 'rb') as f:
            doctor.output_video.save(f"{doctor.id}_output.mp4", File(f), save=True)

        BASE_URL = "https://api.videomaker.digielvestech.in/"
        doctor.output_video_url = f"{BASE_URL}{doctor.output_video.url}"
        doctor.save()

        print(f"Video generated and saved for doctor {doctor.name}")

    except Exception as e:
        print(f"Error generating video for doctor {doctor.name}: {e}")



@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def bulk_upload_doctors(request):
    excel_file = request.FILES.get('file')
    if not excel_file:
        return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        df = pd.read_excel(excel_file)
        required_columns = {'name', 'clinic', 'city', 'specialization', 'state'}

        if not required_columns.issubset(df.columns):
            return Response({'error': f'Missing required columns: {required_columns - set(df.columns)}'},
                            status=status.HTTP_400_BAD_REQUEST)

        created, skipped, errors = 0, 0, []

        for index, row in df.iterrows():
            row_number = index + 2
            name = str(row.get('name', '')).strip()
            clinic = str(row.get('clinic', '')).strip()
            city = str(row.get('city', '')).strip()
            specialization = str(row.get('specialization', '')).strip()
            state = str(row.get('state', '')).strip()

            # Skip empty rows
            if not name or not clinic or not city or not specialization or not state:
                skipped += 1
                errors.append({'row': row_number, 'error': 'Required fields are missing'})
                continue


            designation = str(row.get('designation', '')).strip()
            mobile_number = str(row.get('mobile_number', '')).strip()
            whatsapp_number = str(row.get('whatsapp_number', '')).strip()
            description = str(row.get('description', '')).strip()
            image_path = str(row.get('image_url', '')).strip()
            employee = row.get('emp_id')

            image_file = None
            if image_path and os.path.exists(image_path):
                image_file = File(open(image_path, 'rb'), name=os.path.basename(image_path))

            doctor_data = {
                'name': name,
                'clinic': clinic,
                'city': city,
                'specialization': specialization,
                'state': state,
                'image': image_file,
                'designation': designation,
                'mobile_number': mobile_number,
                'whatsapp_number': whatsapp_number,
                'description': description,
                'employee': employee,
            }

            serializer = DoctorSerializer(data=doctor_data)
            if serializer.is_valid():
                try:
                    doctor = serializer.save()
                    generate_video_for_doctor(doctor)  # ✅ One video per doctor
                    created += 1
                except IntegrityError:
                    skipped += 1
                    errors.append({'row': row_number, 'error': 'Integrity error during save'})
            else:
                skipped += 1
                errors.append({'row': row_number, 'error': serializer.errors})

        return Response({
            'created': created,
            'skipped': skipped,
            'errors': errors
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


class DoctorVideoExportExcelView(APIView):

    def get(self, request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Doctor Videos"

        headers = [
            "Name", "Designation", "Clinic", "City", "State", "Image URL",
            "Specialization", "Mobile Number", "WhatsApp Number", "Description","Template Name",
            "Output Video URL", "Created At", "Employee ID", "Employee Name", "RBM Name"
        ]
        sheet.append(headers)

        doctor_videos = DoctorVideo.objects.all()

        for video in doctor_videos:
            # Get latest DoctorOutputVideo instance
            latest_output = video.doctoroutputvideo_set.order_by('-created_at').first()

            if latest_output and latest_output.video_file:
                output_video_url = request.build_absolute_uri(latest_output.video_file.url)
                template_name = latest_output.template.name if latest_output.template else ""
            elif video.output_video:
                output_video_url = request.build_absolute_uri(video.output_video.url)
                template_name = ""
            else:
                output_video_url = ""
                template_name = ""

            rbm_name = (
                f"{video.employee.rbm.first_name} {video.employee.rbm.last_name}"
                if video.employee and video.employee.rbm else ""
            )

            sheet.append([
                video.name,
                video.designation,
                video.clinic,
                video.city,
                video.state,
                request.build_absolute_uri(video.image.url) if video.image else "",
                video.specialization,
                video.mobile_number,
                video.whatsapp_number,
                video.description,
                template_name,
                output_video_url,  # ✅ Now coming from DoctorOutputVideo
                video.created_at.strftime('%Y-%m-%d %H:%M:%S') if video.created_at else "",
                video.employee.employee_id if video.employee else "",
                f"{video.employee.first_name} {video.employee.last_name}" if video.employee else "",
                rbm_name,
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=doctor_videos.xlsx'
        workbook.save(response)
        return response



class EmployeeExportExcelView(APIView):

    def get(self, request):
        # Create Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Employees"

        # Add headers
        headers = [
            "Employee ID", "First Name", "Last Name", "Email", "Phone",
            "Department", "Date Joined", "User Type", "Status"
        ]
        sheet.append(headers)

        # Populate data
        for emp in Employee.objects.all():
            sheet.append([
                emp.employee_id,
                emp.first_name,
                emp.last_name or "",
                emp.email or "",
                emp.phone or "",
                emp.department or "",
                emp.date_joined.strftime('%Y-%m-%d') if emp.date_joined else "",
                emp.user_type,
                "Active" if emp.status else "Inactive"
            ])

        # Prepare Excel response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=employees.xlsx'
        workbook.save(response)
        return response


@api_view(['GET'])
def total_employee_count(request):
    count = Employee.objects.count()
    return Response({
        "total_employees": count
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def todays_active_employees(request):
    today = timezone.now().date()

    # Get all employees who have login history with login_time today
    employees = Employee.objects.filter(
        login_history__login_time__date=today
    ).distinct()

    # Count of those employees
    count = employees.count()

    data = [{
        'employee_id': emp.employee_id,
        'name': f"{emp.first_name} {emp.last_name}",
        'email': emp.email,
        'department': emp.department,
        'user_type': emp.user_type,
    } for emp in employees]

    return Response({
        'date': str(today),
        'active_employee_count': count,
        'active_employees': data
    })


class TodaysActiveEmployeeExcelExport(APIView):
    def get(self, request):
        today = timezone.now().date()

        # Get employees who logged in today
        employees = Employee.objects.filter(
            login_history__login_time__date=today
        ).distinct()

        # Create Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Today's Active Employees"

        # Add headers
        headers = [
            "Employee ID", "First Name", "Last Name", "Email", "Phone",
            "Department", "Date Joined", "User Type", "Status"
        ]
        sheet.append(headers)

        # Populate data
        for emp in employees:
            sheet.append([
                emp.employee_id,
                emp.first_name,
                emp.last_name or "",
                emp.email or "",
                emp.phone or "",
                emp.department or "",
                emp.date_joined.strftime('%Y-%m-%d') if emp.date_joined else "",
                emp.user_type,
                "Active" if emp.status else "Inactive"
            ])

        # Prepare response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"active_employees_{today}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        workbook.save(response)
        return response
    




@api_view(['GET'])
def doctors_with_output_video_count(request):
    count = DoctorVideo.objects.filter(
        Q(output_video__isnull=False) & ~Q(output_video='')
    ).count()
    return Response({
        "doctor_with_output_video_count": count
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def doctors_with_output_video_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Doctor Videos"

    # Updated header row (added "RBM Name")
    headers = [
        "Name", "Designation", "Clinic", "City", "State", "Image URL",
        "Specialization", "Mobile Number", "WhatsApp Number", "Description",
        "Output Video URL", "Created At", "Employee ID", "Employee Name", "RBM Name"
    ]
    sheet.append(headers)

    # Get doctor videos with output video only
    # doctor_videos = DoctorVideo.objects.filter(
    #     Q(output_video__isnull=False) & ~Q(output_video='')
    # )
    doctor_videos = DoctorVideo.objects.all()

    for video in doctor_videos:
        # Safely get RBM name
        rbm_name = (
            f"{video.employee.rbm.first_name} {video.employee.rbm.last_name}"
            if video.employee and video.employee.rbm else ""
        )

        sheet.append([
            video.name,
            video.designation,
            video.clinic,
            video.city,
            video.state,
            request.build_absolute_uri(video.image.url) if video.image else "",
            video.specialization,
            video.mobile_number,
            video.whatsapp_number,
            video.description,
            request.build_absolute_uri(video.output_video.url) if video.output_video else "",
            video.created_at.strftime('%Y-%m-%d %H:%M:%S') if video.created_at else "",
            video.employee.employee_id if video.employee else "",
            f"{video.employee.first_name} {video.employee.last_name}" if video.employee else "",
            rbm_name,
        ])

    # Return Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=doctor_videos.xlsx'
    workbook.save(response)
    return response


@api_view(['GET'])
def doctors_count(request):
    count = DoctorVideo.objects.filter().count()
    return Response({
        "doctor_count": count
    }, status=status.HTTP_200_OK)



# class VideoTemplateAPIView(APIView):

#     def get(self, request, pk=None):
#         if pk:
#             template = get_object_or_404(VideoTemplates, pk=pk)
#             serializer = VideoTemplatesSerializer(template)
#         else:
#             status_param = request.query_params.get('status')

#             templates = VideoTemplates.objects.all()

#             if status_param is not None:
#                 try:
#                     status_bool = bool(strtobool(status_param))  # Converts "true"/"false" to True/False
#                     templates = templates.filter(status=status_bool)
#                 except ValueError:
#                     return Response({"error": "Invalid status value. Use true or false."}, status=status.HTTP_400_BAD_REQUEST)

#             templates = templates.order_by('-created_at')
#             serializer = VideoTemplatesSerializer(templates, many=True)

#         return Response(serializer.data, status=status.HTTP_200_OK)

#! Added By Prathamesh-

class VideoTemplateAPIView(APIView):

    def get(self, request, pk=None):
        if pk:
            template = get_object_or_404(VideoTemplates, pk=pk)
            print(f"🔍 Fetching template ID: {pk}")
            print(f"🔍 Template custom_text: '{template.custom_text}'")
            print(f"🔍 Template name: '{template.name}'")
            serializer = VideoTemplatesSerializer(template)
        else:
            status_param = request.query_params.get('status')
            template_type = request.query_params.get('template_type', 'video')  # NEW: Add template type filter

            templates = VideoTemplates.objects.filter(template_type=template_type)  # NEW: Filter by type

            # DEBUG: Check what's in database
            print(f"Found {templates.count()} templates of type {template_type}")
            for template in templates[:3]:  # Print first 3
                print(f"Template: {template.name}, Type: {template.template_type}")

            if status_param is not None:
                try:
                    status_bool = bool(strtobool(status_param))
                    templates = templates.filter(status=status_bool)
                except ValueError:
                    return Response({"error": "Invalid status value. Use true or false."}, status=status.HTTP_400_BAD_REQUEST)

            templates = templates.order_by('-created_at')
            serializer = VideoTemplatesSerializer(templates, many=True)
            # print("Serialized data sample:", serializer.data[:1] if serializer.data else "No data")


        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = VideoTemplatesSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        template = get_object_or_404(VideoTemplates, pk=pk)
        serializer = VideoTemplatesSerializer(template, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        template = get_object_or_404(VideoTemplates, pk=pk)
        template.delete()
        return Response({"detail": "Deleted successfully."}, status=status.HTTP_204_NO_CONTENT)



class GenerateDoctorOutputVideoView(APIView):
    def post(self, request):
        doctor_id = request.data.get("doctor_id")
        template_id = request.data.get("template_id")

        if not doctor_id:
            return Response({"error": "doctor_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            doctor = DoctorVideo.objects.get(id=doctor_id)
        except DoctorVideo.DoesNotExist:
            return Response({"error": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)

        # Use default template if not provided
        if template_id:
            try:
                template = VideoTemplates.objects.get(id=template_id)
            except VideoTemplates.DoesNotExist:
                return Response({"error": "Template not found."}, status=status.HTTP_404_NOT_FOUND)
        else:
            template = VideoTemplates.objects.filter(status=True).first()
            if not template:
                return Response({"error": "No default template available."}, status=status.HTTP_400_BAD_REQUEST)

        if not doctor.image:
            return Response({"error": "Doctor does not have an image."}, status=status.HTTP_400_BAD_REQUEST)

        image_path = doctor.image.path

        # Prepare output path
        random_key = uuid.uuid4().hex[:8]
        output_filename = f"{doctor.id}_{template.id}_{random_key}_output.mp4"

        output_dir = os.path.join(settings.MEDIA_ROOT, "output", str(doctor.employee.id), str(doctor.id))
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, output_filename)

        print(f"🔍 Template video path: {template.template_video.path}")
        print(f"🔍 Template video file exists: {os.path.exists(template.template_video.path)}")
        print(f"🔍 Output directory: {output_dir}")
        print(f"🔍 Output path: {output_path}")
        print(f"🔍 Image path: {image_path}")
        print(f"🔍 Image file exists: {os.path.exists(image_path)}")

        try:
            self.generate_custom_video(
                main_video_path=template.template_video.path,
                image_path=image_path,
                name=doctor.name,
                clinic=doctor.clinic,
                city=doctor.city,
                specialization_key=doctor.specialization_key,
                state=doctor.state,
                output_path=output_path,
                resolution=template.resolution,
                base_x=template.base_x_axis,
                base_y=template.base_y_axis,
                line_spacing=template.line_spacing,
                overlay_x=template.overlay_x,
                overlay_y=template.overlay_y,

            )

            relative_path = os.path.relpath(output_path, settings.MEDIA_ROOT)

            output_video = DoctorOutputVideo.objects.create(
                doctor=doctor,
                template=template,
                video_file=relative_path
            )

            # with open(output_path, 'rb') as f:
            #     output_video = DoctorOutputVideo.objects.create(
            #         doctor=doctor,
            #         template=template,
            #         video_file=File(f, name=output_filename)
            #     )

            serializer = DoctorOutputVideoSerializer(output_video)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            print(f"Video generation failed: {e}")
            return Response({"error": "Video generation failed.", "details": str(e)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def generate_custom_video(self, main_video_path, image_path, name, clinic, city, specialization_key, state, output_path,
                          resolution="415x410", base_x="(main_w/2)-160", base_y="(main_h/2)-60", line_spacing="60",overlay_x="350",
                          overlay_y="70"):
        temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
        os.makedirs(temp_dir, exist_ok=True)

        fps = 30
        fade_duration = 3   
        slots = [(2, 6), (65, 70)]

        temp_videos = []
        for i, (start, end) in enumerate(slots):
            duration = end - start
            total_frames = duration * fps

            zoom_effect = f"zoompan=z='1+0.00003*in':x='(iw/2)-(iw/zoom/2)':y='(ih/2)-(ih/zoom/2)':d={total_frames}:s={resolution}:fps={fps}"
            fade_effect = f"format=rgba,fade=t=in:st=0:d={fade_duration}:alpha=1,fade=t=out:st={duration-fade_duration}:d={fade_duration}:alpha=1"
            vf = f"scale={resolution},{zoom_effect},{fade_effect}"

            temp_video = os.path.join(temp_dir, f"temp_image_vid_{i}.mp4")
            # Create a simple test command first

            
            subprocess.run([
                r'C:\ffmpeg\bin\ffmpeg.exe', "-loop", "1", "-i", image_path,
                "-vf", vf, "-t", str(duration), "-y", temp_video
            ], check=True)

            temp_videos.append((temp_video, start, end))

        # === Drawtext settings ===
        text_lines = [name, specialization_key, clinic, city, state]
        font = "RobotoSlab-Medium.ttf"

        text_filters = []
        for start, end in slots:
            alpha_expr = f"if(lt(t\\,{start}+3),(t-{start})/3,if(lt(t\\,{end}-3),1,({end}-t)/3))"
            for j, text in enumerate(text_lines):
                # y_pos = f"{base_y}+{j}*{line_spacing}"
                # drawtext = (
                #     f"drawtext=text='{text}':fontfile='{font}':fontcolor=black:fontsize=40:"
                #     f"x={base_x}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
                # )



                offset = 132  # adjust this value to move text further up or down
                y_pos = f"(main_h - ({len(text_lines)}*{line_spacing}) + {j}*{line_spacing} - {offset})"

                x_pos= "(main_w/2)-300"
               
                drawtext = (
                    f"drawtext=text='{text}':fontfile='{font}':fontcolor=black:fontsize=40:"
                    f"x={x_pos}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
                )

                text_filters.append(drawtext)

        overlay_x1 = f"(main_w-overlay_w)/2-{int(overlay_x)}"
        overlay_y1 = f"(main_h-overlay_h)/2+{int(overlay_y)}"

        filter_complex = (
            f"[0:v][1:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{slots[0][0]},{slots[0][1]})'[v1];"
            f"[v1][2:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{slots[1][0]},{slots[1][1]})'[v2];"
            f"[v2]{','.join(text_filters)}[v]"
        )

        
        cmd = [
            "ffmpeg", "-i", main_video_path,
            "-i", temp_videos[0][0],
            "-i", temp_videos[1][0],
            "-filter_complex", filter_complex,
            "-map", "[v]",
            "-map", "0:a?",
            "-c:v", "libx264", "-c:a", "copy", "-y", output_path
        ]

        

        print(f"🔍 Running FFmpeg command: {' '.join(cmd[:3])}...")
        subprocess.run(cmd, check=True)

        for temp_video, _, _ in temp_videos:
            os.remove(temp_video)

    def get(self, request):
        doctor_id = request.query_params.get("doctor_id")
        employee_id = request.query_params.get("employee_id")

        videos = DoctorOutputVideo.objects.all().order_by('-id')  # Latest created first

        if doctor_id:
            videos = videos.filter(doctor_id=doctor_id)

        if employee_id:
            videos = videos.filter(doctor__employee_id=employee_id)

        paginator = Pagination_class()
        paginated_videos = paginator.paginate_queryset(videos, request)

        serializer = DoctorOutputVideoSerializer(paginated_videos, many=True)
        return paginator.get_paginated_response(serializer.data)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def update_employees_from_excel(request):
    excel_file = request.FILES.get('file')
    
    if not excel_file:
        return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        df = pd.read_excel(excel_file)

        updated, not_found = 0, 0
        for _, row in df.iterrows():
            employee_id = str(row['id']).strip()
            department = str(row['department']).strip()
            city = str(row['city']).strip().title()

            try:
                employee = Employee.objects.get(employee_id=employee_id)
                employee.department = department
                employee.city = city
                employee.save()
                updated += 1
            except Employee.DoesNotExist:
                not_found += 1

        return Response({
            'status': 'success',
            'updated': updated,
            'not_found': not_found
        })

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class TemplateWiseVideoCountView(APIView):
    def get(self, request):
        template_counts = DoctorOutputVideo.objects.values(
            "template__id", "template__name"
        ).annotate(
            video_count=Count("id")
        ).order_by("-video_count")

        data = [
            {
                "template_id": item["template__id"],
                "template_name": item["template__name"],
                "video_count": item["video_count"]
            }
            for item in template_counts if item["template__id"] is not None
        ]

        return Response(data, status=status.HTTP_200_OK)



#! ADDED BY Prathamesh

from distutils.util import strtobool 


class ImageTemplateAPIView(APIView):
    """Handle image templates separately from video templates"""
    
    def get(self, request, pk=None):
        if pk:
            template = get_object_or_404(VideoTemplates, pk=pk, template_type='image')
            serializer = ImageTemplateSerializer(template, context={'request': request})
        else:
            status_param = request.query_params.get('status')
            
            templates = VideoTemplates.objects.filter(template_type='image')
            
            if status_param is not None:
                try:
                    status_bool = bool(strtobool(status_param))
                    templates = templates.filter(status=status_bool)
                except ValueError:
                    return Response({"error": "Invalid status value. Use true or false."}, 
                                  status=status.HTTP_400_BAD_REQUEST)
            
            templates = templates.order_by('-created_at')
            serializer = ImageTemplateSerializer(templates, many=True, context={'request': request})
        
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def post(self, request):
        serializer = ImageTemplateSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self, request, pk):
        template = get_object_or_404(VideoTemplates, pk=pk, template_type='image')
        serializer = ImageTemplateSerializer(template, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        template = get_object_or_404(VideoTemplates, pk=pk, template_type='image')
        template.delete()
        return Response({"detail": "Template deleted successfully."}, status=status.HTTP_204_NO_CONTENT)


# class GenerateImageContentView(APIView):
#     """Generate image with text overlay"""
#     def post(self, request):
#         print(f"🔍 DJANGO: Received POST data: {request.data}")
#         print(f"🔍 DJANGO: Doctor data received: {request.data.get('doctor_data')}")
#         template_id = request.data.get("template_id")
#         doctor_id = request.data.get("doctor_id")  # Existing doctor scenario
#         mobile = request.data.get("mobile")        # New doctor scenario
#         content_data = request.data.get("content_data", {})
#         doctor_data = request.data.get("doctor_data")
        
#         print(f"🔍 DJANGO: Extracted doctor_data: {doctor_data}")
#         print(f"🔍 DJANGO: Mobile: {mobile}")

#         if not template_id:
#             return Response({"error": "template_id is required."}, status=status.HTTP_400_BAD_REQUEST)
        
#         # Must have either doctor_id OR mobile OR doctor_data
#         if not doctor_id and not mobile and not doctor_data:
#             return Response({"error": "Either doctor_id, mobile, or doctor_data is required."}, 
#                         status=status.HTTP_400_BAD_REQUEST)
        
#         try:
#             template = VideoTemplates.objects.get(id=template_id, template_type='image')
#         except VideoTemplates.DoesNotExist:
#             return Response({"error": "Image template not found."}, status=status.HTTP_404_NOT_FOUND)
        
#         # Scenario 1: Existing doctor by ID
#         if doctor_id:
#             # Try DoctorVideo first
#             doctor_video = DoctorVideo.objects.filter(id=doctor_id).first()
            
#             if not doctor_video:
#                 # Try Doctor model and convert to DoctorVideo
#                 doctor = Doctor.objects.filter(id=doctor_id).first()
#                 if doctor:
#                     # Check if DoctorVideo already exists for this mobile
#                     doctor_video = DoctorVideo.objects.filter(mobile_number=doctor.mobile_number).first()
                    
#                     if not doctor_video:
#                         # Create DoctorVideo from Doctor
#                         doctor_video = DoctorVideo.objects.create(
#                             name=doctor.name,
#                             designation=doctor.designation,
#                             clinic=doctor.clinic,
#                             city=doctor.city,
#                             state=doctor.state,
#                             specialization=doctor.specialization,
#                             specialization_key=doctor.specialization,
#                             mobile_number=doctor.mobile_number,
#                             whatsapp_number=doctor.whatsapp_number,
#                             description=doctor.description,
#                             employee=doctor.employee,
#                             image=doctor.image
#                         )
            
#             if not doctor_video:
#                 return Response({"error": "Doctor not found in either model."}, status=status.HTTP_404_NOT_FOUND)
            
#             is_new_doctor = False
        
#         # Scenario 2: Doctor by mobile
#         # Scenario 2: Doctor by mobile
#         elif mobile:
#             print(f"🔍 ENTERING SCENARIO 2 - Mobile: {mobile}")
#             name = request.data.get("name")
#             if not name:
#                 return Response({"error": "name is required when using mobile."}, 
#                             status=status.HTTP_400_BAD_REQUEST)
            
#             print(f"🔍 Name: {name}")
            
#             # Search in DoctorVideo first
#             doctor_video = DoctorVideo.objects.filter(mobile_number=mobile).first()
#             print(f"🔍 Existing doctor_video found: {doctor_video}")
            
#             if not doctor_video:
#                 # Search in Doctor model and convert
#                 doctor = Doctor.objects.filter(mobile_number=mobile).first()
#                 if doctor:
#                     # Create DoctorVideo from existing Doctor
#                     doctor_video = DoctorVideo.objects.create(
#                         name=doctor.name,
#                         designation=doctor.designation,
#                         clinic=doctor.clinic,
#                         city=doctor.city,
#                         state=doctor.state,
#                         specialization=doctor.specialization,
#                         specialization_key=doctor.specialization,
#                         mobile_number=doctor.mobile_number,
#                         whatsapp_number=doctor.whatsapp_number,
#                         description=doctor.description,
#                         employee=doctor.employee,
#                         image=doctor.image
#                     )
#                     is_new_doctor = False
#                 else:
#                     # Create completely new DoctorVideo
#                     try:
#                         employee_id = request.data.get("employee_id")
#                         if employee_id:
#                             employee = Employee.objects.get(employee_id=employee_id)
#                         else:
#                             employee = Employee.objects.first()  # Use first available employee
#                             if not employee:
#                                 return Response({"error": "No employees found in system."}, status=status.HTTP_404_NOT_FOUND)
#                     except Employee.DoesNotExist:
#                         return Response({"error": f"Employee {employee_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                    

#                     doctor_data = request.data.get("doctor_data", {})

#                     print(f"🔍 CREATING NEW DOCTOR with these parameters:")
#                     print(f"🔍 name: '{name}'")
#                     print(f"🔍 clinic from doctor_data: '{doctor_data.get('clinic', 'NOT_FOUND')}'")
#                     print(f"🔍 city from doctor_data: '{doctor_data.get('city', 'NOT_FOUND')}'")
#                     print(f"🔍 specialization from doctor_data: '{doctor_data.get('specialization', 'NOT_FOUND')}'")
#                     print(f"🔍 state from doctor_data: '{doctor_data.get('state', 'NOT_FOUND')}'")
#                     print(f"🔍 doctor_data variable: {doctor_data}")

#                     # Try to create with explicit values first
#                     clinic_value = doctor_data.get("clinic", "") if doctor_data else ""
#                     city_value = doctor_data.get("city", "") if doctor_data else ""
#                     specialization_value = doctor_data.get("specialization", "") if doctor_data else ""
#                     state_value = doctor_data.get("state", "") if doctor_data else ""

#                     print(f"🔍 Final values to save:")
#                     print(f"🔍 clinic_value: '{clinic_value}'")
#                     print(f"🔍 city_value: '{city_value}'")
#                     print(f"🔍 specialization_value: '{specialization_value}'")
#                     print(f"🔍 state_value: '{state_value}'")



#                     doctor_video = DoctorVideo.objects.create(
#                         name=name,
#                         clinic=clinic_value,
#                         city=city_value,
#                         specialization=specialization_value,
#                         specialization_key=specialization_value,
#                         state=state_value,
#                         mobile_number=mobile,
#                         whatsapp_number=mobile,
#                         description="",
#                         designation="",
#                         employee=employee
#                     )
#                     print(f"🔍 CREATED doctor with ID: {doctor_video.id}")
#                     print(f"🔍 Immediately after create - clinic: '{doctor_video.clinic}'")
#                     print(f"🔍 Immediately after create - city: '{doctor_video.city}'")
#                     print(f"🔍 Immediately after create - specialization: '{doctor_video.specialization}'")
#                     print(f"🔍 Immediately after create - state: '{doctor_video.state}'")
#                     is_new_doctor = True
#             else:
#                 is_new_doctor = False
        
#         # Scenario 3: Doctor data object
#         elif doctor_data:
#             mobile = doctor_data.get("mobile")
#             name = doctor_data.get("name")
            
#             if not mobile or not name:
#                 return Response({"error": "mobile and name are required in doctor_data."}, 
#                             status=status.HTTP_400_BAD_REQUEST)
            
#             # Search in DoctorVideo first
#             doctor_video = DoctorVideo.objects.filter(mobile_number=mobile).first()
            
#             if not doctor_video:
#                 # Search in Doctor model and convert
#                 doctor = Doctor.objects.filter(mobile_number=mobile).first()
#                 if doctor:
#                     # Create DoctorVideo from existing Doctor
#                     doctor_video = DoctorVideo.objects.create(
#                         name=doctor.name,
#                         designation=doctor.designation,
#                         clinic=doctor.clinic,
#                         city=doctor.city,
#                         state=doctor.state,
#                         specialization=doctor.specialization,
#                         specialization_key=doctor.specialization,
#                         mobile_number=doctor.mobile_number,
#                         whatsapp_number=doctor.whatsapp_number,
#                         description=doctor.description,
#                         employee=doctor.employee,
#                         image=doctor.image
#                     )
#                     is_new_doctor = False
#                 else:
#                     # Create completely new DoctorVideo
# # Create completely new DoctorVideo
#                     try:
#                         employee_id = request.data.get("employee_id")
#                         if employee_id:
#                             employee = Employee.objects.get(employee_id=employee_id)
#                         else:
#                             employee = Employee.objects.first()
#                             if not employee:
#                                 return Response({"error": "No employees found in system."}, status=status.HTTP_404_NOT_FOUND)
#                     except Employee.DoesNotExist:
#                         return Response({"error": f"Employee {employee_id} not found."}, status=status.HTTP_404_NOT_FOUND)

#                     # Use doctor_data if available, otherwise get from request.data directly
#                     doctor_info = request.data.get("doctor_data", {})
#                     if not doctor_info:
#                         # Fallback to direct request data
#                         doctor_info = {
#                             'clinic': request.data.get("clinic", ""),
#                             'city': request.data.get("city", ""),
#                             'specialization': request.data.get("specialization", ""),
#                             'state': request.data.get("state", "")
#                         }

#                     print(f"🔍 Using doctor_info: {doctor_info}")

#                     doctor_video = DoctorVideo.objects.create(
#                         name=name,
#                         clinic=doctor_info.get("clinic", ""),
#                         city=doctor_info.get("city", ""),
#                         specialization=doctor_info.get("specialization", ""),
#                         specialization_key=doctor_info.get("specialization", ""),
#                         state=doctor_info.get("state", ""),
#                         mobile_number=mobile,
#                         whatsapp_number=mobile,
#                         description="",
#                         designation="",
#                         employee=employee
#                     )
#                     is_new_doctor = True
#             else:
#                 is_new_doctor = False

#         if not template.template_image:
#             return Response({"error": "Template does not have an image."}, 
#                         status=status.HTTP_400_BAD_REQUEST)
        
#         try:
#             output_path = self.generate_image_with_text(template, content_data, doctor_video)
            
#             # Create ImageContent record (now always with DoctorVideo)
#             image_content = ImageContent.objects.create(
#                 template=template,
#                 doctor=doctor_video,  # Always DoctorVideo now!
#                 content_data=content_data
#             )
            
#             # Save generated image
#             with open(output_path, 'rb') as f:
#                 image_content.output_image.save(
#                     f"generated_{image_content.id}.png", 
#                     File(f), 
#                     save=True
#                 )
            
#             os.remove(output_path)
            
#             serializer = ImageContentSerializer(image_content, context={'request': request})
#             response_data = serializer.data
#             response_data['doctor_info'] = {
#                 'id': doctor_video.id,
#                 'name': doctor_video.name,
#                 'clinic': doctor_video.clinic,
#                 'mobile': doctor_video.mobile_number,
#                 'is_new_doctor': is_new_doctor
#             }
            
#             return Response(response_data, status=status.HTTP_201_CREATED)
            
#         except Exception as e:
#             logger.error(f"Image generation failed: {e}")
#             return Response({"error": "Image generation failed.", "details": str(e)},
#                         status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    
    
# #     def generate_image_with_text(self, template, content_data, doctor):
# #         """Generate image with text overlay using PIL"""
        
# #         # Load template image
# #         template_image = Image.open(template.template_image.path)
# #         draw = ImageDraw.Draw(template_image)
        
# #         # Get text positions from template
# #         positions = template.text_positions or {}
# #         print(f"Template positions: {positions}")  # DEBUG
# #         print(f"Template custom text: {template.custom_text}")  # DEBUG
        
# #         # Try to load a better font
# # # Font mapping for PIL
# #         FONT_MAP = {
# #             'Arial': 'arial.ttf',
# #             'Times New Roman': 'times.ttf', 
# #             'Helvetica': 'arial.ttf',  # Fallback to arial
# #             'Georgia': 'georgia.ttf',
# #             'Verdana': 'verdana.ttf',
# #             'Impact': 'impact.ttf',
# #             'Comic Sans MS': 'comic.ttf'
# #         }

# #         def get_font(font_family, font_size):
# #             """Get PIL font with fallback"""
# #             font_file = FONT_MAP.get(font_family, 'arial.ttf')
# #             try:
# #                 return ImageFont.truetype(font_file, font_size)
# #             except:
# #                 try:
# #                     return ImageFont.truetype("arial.ttf", font_size)
# #                 except:
# #                     return ImageFont.load_default()
        
# #         # Prepare all text data
# #         # Prepare all text data
# #         all_text_data = {
# #             # Doctor info from database
# #             'name': doctor.name,
# #             'clinic': doctor.clinic,
# #             'city': doctor.city,
# #             'specialization': doctor.specialization,
# #             'state': doctor.state,
# #             'mobile': doctor.mobile_number,
            
# #             # Override with dynamic content_data if provided (from Profile form)
# #             'name': content_data.get('doctor_name', doctor.name),
# #             'clinic': content_data.get('doctor_clinic', doctor.clinic),
# #             'city': content_data.get('doctor_city', doctor.city),
# #             'specialization': content_data.get('doctor_specialization', doctor.specialization),
# #             'state': content_data.get('doctor_state', doctor.state),
            
# #             # Template's custom text (use template's custom_text, not hardcoded message)
# #             # Template's custom text ONLY - no static fallbacks
# #             # 'customText': template.custom_text or content_data.get('custom_text', ''),
# #             #! started working
# #             'customText': template.custom_text or '', 
            
# #         }
# #         print(f"🔍 Template custom_text: '{template.custom_text}'")
# #         print(f"🔍 Content data custom_text: '{content_data.get('custom_text', '')}'")
# #         print(f"🔍 Final customText value: '{template.custom_text or content_data.get('custom_text', '')}'")
# #         print(f"🔍 All text data: {all_text_data}")
                
# #         # Add text for each field with proper styling
# #         for field_name, text_value in all_text_data.items():
# #             if field_name in positions and text_value:
# #                 pos = positions[field_name]
                
# #                 # Get styling from position data
# # # Get styling from position data
# #                 font_size = int(pos.get('fontSize', 40))
# #                 color = pos.get('color', 'black')
# #                 font_weight = pos.get('fontWeight', 'normal')
# #                 font_family = pos.get('fontFamily', 'Arial')

# #                 # Load font with proper styling
# #                 styled_font = get_font(font_family, font_size)

# #                 # Handle bold font weight
# #                 if font_weight == 'bold':
# #                     try:
# #                         # Try to get bold version
# #                         bold_font_file = FONT_MAP.get(font_family, 'arial.ttf').replace('.ttf', 'bd.ttf')
# #                         styled_font = ImageFont.truetype(bold_font_file, font_size)
# #                     except:
# #                         # Fallback to regular font
# #                         styled_font = get_font(font_family, font_size)
                
# #                 # print(f"Rendering {field_name}: '{text_value}' at ({pos['x']}, {pos['y']}) with color {color}")  # DEBUG
                
# #                 # draw.text(
# #                 #     (int(pos['x']), int(pos['y'])), 
# #                 #     str(text_value), 
# #                 #     fill=color, 
# #                 #     font=styled_font
# #                 # )

# #                 # Handle text shadow
# #                 text_shadow = pos.get('textShadow', 'none')
# #                 shadow_info = parse_css_shadow(text_shadow)

# #                 if shadow_info:
# #                     # Draw shadow first
# #                     draw.text(
# #                         (int(pos['x']) + shadow_info['offset_x'], int(pos['y']) + shadow_info['offset_y']), 
# #                         str(text_value), 
# #                         fill=shadow_info['color'], 
# #                         font=styled_font
# #                     )

# #                 # Draw main text on top
# #                 draw.text(
# #                     (int(pos['x']), int(pos['y'])), 
# #                     str(text_value), 
# #                     fill=color, 
# #                     font=styled_font
# #                 )

# #             elif text_value:
# #                     # Fallback for unmapped fields
# #                     print(f"Warning: No position found for field '{field_name}' with value '{text_value}'")
                    
# #                     # Default positions for unmapped fields
# #                     default_positions = {
# #                         'user_message': {'x': 100, 'y': 400, 'fontSize': 30, 'color': 'green'},
# #                     }
# #                     if field_name in default_positions:
# #                         pos = default_positions[field_name]
# #                         try:
# #                             fallback_font = ImageFont.truetype("arial.ttf", int(pos['fontSize']))
# #                         except:
# #                             fallback_font = ImageFont.load_default()
                            
# #                         draw.text(
# #                             (int(pos['x']), int(pos['y'])), 
# #                             str(text_value), 
# #                             fill=pos['color'],
# #                             font=fallback_font
# #                         )
# #                 # Continue without doctor image if there's an error
# #         # Save to temp location
# #        # Handle doctor image overlay if enabled
# #         # image_settings = template.text_positions.get('imageSettings') if template.text_positions else None
# #         # Handle doctor image overlay if enabled - ENHANCED
# #         print(f"🔍 ===== DOCTOR IMAGE DEBUG =====")
# #         print(f"🔍 Template text_positions: {template.text_positions}")
# #         print(f"🔍 Content data: {content_data}")

# #         image_settings = None

# #         # Priority 1: Check content_data for image settings (from live preview)
# #         if content_data and 'imageSettings' in content_data:
# #             image_settings = content_data['imageSettings']
# #             print(f"🔍 Using image settings from content_data: {image_settings}")

# #         # Priority 2: Check template for stored image settings
# #         elif template.text_positions and 'imageSettings' in template.text_positions:
# #             image_settings = template.text_positions['imageSettings']
# #             print(f"🔍 Using image settings from template: {image_settings}")

# #         print(f"🔍 Final image settings: {image_settings}")

# #         # DEBUG: Check doctor image status
# #         print(f"🔍 Doctor object: {doctor}")
# #         print(f"🔍 Doctor ID: {doctor.id}")
# #         print(f"🔍 Doctor name: {doctor.name}")
# #         print(f"🔍 Doctor.image field: {doctor.image}")
# #         print(f"🔍 Doctor.image exists: {bool(doctor.image)}")
# #         print(f"🔍 Doctor.image type: {type(doctor.image)}")

# #         if doctor.image:
# #             print(f"🔍 Doctor.image.name: {doctor.image.name}")
# #             print(f"🔍 Doctor.image.path: {doctor.image.path}")
# #             try:
# #                 print(f"🔍 Image file exists on disk: {os.path.exists(doctor.image.path)}")
# #             except Exception as e:
# #                 print(f"🔍 Error checking image path: {e}")
# #         else:
# #             print(f"🔍 ❌ Doctor has no image file - this is why image overlay is not working!")
        
# #         # if image_settings and image_settings.get('enabled', False) and doctor.image:
# #         # FORCE ENABLE FOR TESTING
# #         # FORCE ENABLE FOR TESTING - WITH DEFAULT IMAGE
# #         # FORCE ENABLE FOR TESTING - WITH DEFAULT IMAGE
# #         if doctor.image or True:  # Always try to add image for testing
# #             if not image_settings:
# #                 image_settings = {
# #                     'enabled': True,
# #                     'x': 400,
# #                     'y': 50,
# #                     'width': 150,
# #                     'height': 150,
# #                     'fit': 'cover',
# #                     'borderRadius': 50,
# #                     'opacity': 100
# #                 }
# #             else:
# #                 image_settings['enabled'] = True
                
# #             print(f"🔍 FORCE ENABLED image settings: {image_settings}")
            
# #             try:
# #                 # Get positioning settings FIRST
# #                 img_x = int(image_settings.get('x', 400))
# #                 img_y = int(image_settings.get('y', 50))
# #                 img_width = int(image_settings.get('width', 150))
# #                 img_height = int(image_settings.get('height', 150))
# #                 img_fit = image_settings.get('fit', 'cover')
# #                 border_radius = int(image_settings.get('borderRadius', 0))
# #                 opacity = int(image_settings.get('opacity', 100))
                
# #                 # Load or create doctor image
# #                 if not doctor.image:
# #                     print(f"🔍 Doctor has no image - creating placeholder")
# #                     # Create a simple colored placeholder image
# #                     doctor_img = Image.new('RGB', (img_width, img_height), color='lightblue')
# #                     draw = ImageDraw.Draw(doctor_img)
                    
# #                     # Add a simple doctor icon/text
# #                     try:
# #                         # Try to use a larger font size for the emoji
# #                         font_size = min(img_width, img_height) // 3
# #                         doctor_img = Image.new('RGB', (img_width, img_height), color='#4A90E2')
# #                         draw = ImageDraw.Draw(doctor_img)
# #                         draw.text((img_width//2, img_height//2), "👨‍⚕️", fill='white', anchor="mm")
# #                     except:
# #                         # Fallback to simple text
# #                         draw.text((img_width//2, img_height//2), "DR", fill='white', anchor="mm")
                    
# #                     print(f"🔍 Created placeholder image {img_width}x{img_height}")
# #                 else:
# #                     # Use actual doctor image
# #                     doctor_img = Image.open(doctor.image.path)
# #                     print(f"🔍 Loaded actual doctor image")
                    
# #                     # Resize doctor image based on fit mode
# #                     if img_fit == 'cover':
# #                         doctor_img = doctor_img.resize((img_width, img_height), Image.Resampling.LANCZOS)
# #                     elif img_fit == 'contain':
# #                         doctor_img.thumbnail((img_width, img_height), Image.Resampling.LANCZOS)
# #                     elif img_fit == 'stretch':
# #                         doctor_img = doctor_img.resize((img_width, img_height), Image.Resampling.LANCZOS)
                
# #                 # Apply border radius if specified
# #                 if border_radius > 0:
# #                     mask = Image.new('L', (doctor_img.width, doctor_img.height), 0)
# #                     mask_draw = ImageDraw.Draw(mask)
# #                     actual_radius = min(doctor_img.width, doctor_img.height) * border_radius // 200
# #                     mask_draw.rounded_rectangle(
# #                         [(0, 0), (doctor_img.width, doctor_img.height)],
# #                         radius=actual_radius,
# #                         fill=255
# #                     )
                    
# #                     # Convert to RGBA and apply mask
# #                     if doctor_img.mode != 'RGBA':
# #                         doctor_img = doctor_img.convert('RGBA')
# #                     doctor_img.putalpha(mask)
                
# #                 # Apply opacity
# #                 if opacity < 100:
# #                     if doctor_img.mode != 'RGBA':
# #                         doctor_img = doctor_img.convert('RGBA')
# #                     alpha = doctor_img.split()[-1]
# #                     alpha = alpha.point(lambda p: int(p * opacity / 100))
# #                     doctor_img.putalpha(alpha)
                
# #                 # Ensure template image is in RGBA mode for compositing
# #                 if template_image.mode != 'RGBA':
# #                     template_image = template_image.convert('RGBA')
                
# #                 # Paste doctor image onto template
# #                 if doctor_img.mode == 'RGBA':
# #                     template_image.paste(doctor_img, (img_x, img_y), doctor_img)
# #                 else:
# #                     template_image.paste(doctor_img, (img_x, img_y))
                
# #                 print(f"✅ Doctor image composited at ({img_x}, {img_y}) with size {img_width}x{img_height}")
                
# #             except Exception as e:
# #                 print(f"❌ Error compositing doctor image: {e}")
# #                 import traceback
# #                 traceback.print_exc()
# #             if not image_settings:
# #                 image_settings = {
# #                     'enabled': True,
# #                     'x': 400,
# #                     'y': 50,
# #                     'width': 150,
# #                     'height': 150,
# #                     'fit': 'cover',
# #                     'borderRadius': 50,
# #                     'opacity': 100
# #                 }
# #             else:
# #                 image_settings['enabled'] = True
        
# #             print(f"🔍 FORCE ENABLED image settings: {image_settings}")
# #             try:
# #                 # Load doctor image
# #                 doctor_img = Image.open(doctor.image.path)
                
# #                 # Get positioning settings
# #                 img_x = int(image_settings.get('x', 400))
# #                 img_y = int(image_settings.get('y', 50))
# #                 img_width = int(image_settings.get('width', 150))
# #                 img_height = int(image_settings.get('height', 150))
# #                 img_fit = image_settings.get('fit', 'cover')
# #                 border_radius = int(image_settings.get('borderRadius', 0))
# #                 opacity = int(image_settings.get('opacity', 100))
                
# #                 # Resize doctor image based on fit mode
# #                 if img_fit == 'cover':
# #                     # Crop to fit exactly
# #                     doctor_img = doctor_img.resize((img_width, img_height), Image.Resampling.LANCZOS)
# #                 elif img_fit == 'contain':
# #                     # Fit within bounds maintaining aspect ratio
# #                     doctor_img.thumbnail((img_width, img_height), Image.Resampling.LANCZOS)
# #                 elif img_fit == 'stretch':
# #                     # Stretch to exact dimensions
# #                     doctor_img = doctor_img.resize((img_width, img_height), Image.Resampling.LANCZOS)
                
# #                 # Apply border radius if specified
# #                 if border_radius > 0:
# #                     # Create rounded corners
# #                     mask = Image.new('L', (doctor_img.width, doctor_img.height), 0)
# #                     mask_draw = ImageDraw.Draw(mask)
                    
# #                     # Calculate actual radius (border_radius is percentage)
# #                     actual_radius = min(doctor_img.width, doctor_img.height) * border_radius // 200
                    
# #                     mask_draw.rounded_rectangle(
# #                         [(0, 0), (doctor_img.width, doctor_img.height)],
# #                         radius=actual_radius,
# #                         fill=255
# #                     )
                    
# #                     # Apply mask
# #                     doctor_img.putalpha(mask)
                
# #                 # Apply opacity
# #                 if opacity < 100:
# #                     if doctor_img.mode != 'RGBA':
# #                         doctor_img = doctor_img.convert('RGBA')
                    
# #                     # Create alpha channel based on opacity
# #                     alpha = doctor_img.split()[-1]
# #                     alpha = alpha.point(lambda p: int(p * opacity / 100))
# #                     doctor_img.putalpha(alpha)
                
# #                 # Ensure template image is in RGBA mode for compositing
# #                 if template_image.mode != 'RGBA':
# #                     template_image = template_image.convert('RGBA')
                
# #                 # Paste doctor image onto template
# #                 if doctor_img.mode == 'RGBA':
# #                     template_image.paste(doctor_img, (img_x, img_y), doctor_img)
# #                 else:
# #                     template_image.paste(doctor_img, (img_x, img_y))
                
# #                 print(f"✅ Doctor image composited at ({img_x}, {img_y}) with size {img_width}x{img_height}")
                
# #             except Exception as e:
# #                 print(f"❌ Error compositing doctor image: {e}")
# #                 # Continue without doctor image if there's an error

# #         # Save to temp location
# #         output_dir = os.path.join(settings.MEDIA_ROOT, "temp")
# #         os.makedirs(output_dir, exist_ok=True)
# #         output_path = os.path.join(output_dir, f"temp_image_{uuid.uuid4().hex}.png")
# #         template_image.save(output_path)
        
# #         return output_path
    
            
#     def generate_image_with_text(self, template, content_data, doctor):
#         """Generate image with text overlay using PIL"""
        
#         # Load template image
#         template_image = Image.open(template.template_image.path)
#         draw = ImageDraw.Draw(template_image)
        
#         # Get text positions from template
#         positions = template.text_positions or {}
#         print(f"Template positions: {positions}")  # DEBUG
#         print(f"Template custom text: {template.custom_text}")  # DEBUG
        
#         # Font mapping for PIL
#         FONT_MAP = {
#             'Arial': 'arial.ttf',
#             'Times New Roman': 'times.ttf', 
#             'Helvetica': 'arial.ttf',  # Fallback to arial
#             'Georgia': 'georgia.ttf',
#             'Verdana': 'verdana.ttf',
#             'Impact': 'impact.ttf',
#             'Comic Sans MS': 'comic.ttf'
#         }

#         def get_font(font_family, font_size):
#             """Get PIL font with fallback"""
#             font_file = FONT_MAP.get(font_family, 'arial.ttf')
#             try:
#                 return ImageFont.truetype(font_file, font_size)
#             except:
#                 try:
#                     return ImageFont.truetype("arial.ttf", font_size)
#                 except:
#                     return ImageFont.load_default()
        
#         # Prepare all text data
#         all_text_data = {
#             # Override with dynamic content_data if provided (from Profile form)
#             'name': content_data.get('doctor_name', doctor.name),
#             'clinic': content_data.get('doctor_clinic', doctor.clinic),
#             'city': content_data.get('doctor_city', doctor.city),
#             'specialization': content_data.get('doctor_specialization', doctor.specialization),
#             'state': content_data.get('doctor_state', doctor.state),
#             'mobile': doctor.mobile_number,
            
#             # Template's custom text
#             'customText': template.custom_text or '', 
#         }
#         print(f"🔍 Template custom_text: '{template.custom_text}'")
#         print(f"🔍 Content data custom_text: '{content_data.get('custom_text', '')}'")
#         print(f"🔍 All text data: {all_text_data}")
                
#         # Add text for each field with proper styling
#         for field_name, text_value in all_text_data.items():
#             if field_name in positions and text_value:
#                 pos = positions[field_name]
                
#                 # Get styling from position data
#                 font_size = int(pos.get('fontSize', 40))
#                 color = pos.get('color', 'black')
#                 font_weight = pos.get('fontWeight', 'normal')
#                 font_family = pos.get('fontFamily', 'Arial')

#                 # Load font with proper styling
#                 styled_font = get_font(font_family, font_size)

#                 # Handle bold font weight
#                 if font_weight == 'bold':
#                     try:
#                         # Try to get bold version
#                         bold_font_file = FONT_MAP.get(font_family, 'arial.ttf').replace('.ttf', 'bd.ttf')
#                         styled_font = ImageFont.truetype(bold_font_file, font_size)
#                     except:
#                         # Fallback to regular font
#                         styled_font = get_font(font_family, font_size)

#                 # Handle text shadow
#                 text_shadow = pos.get('textShadow', 'none')
#                 shadow_info = parse_css_shadow(text_shadow)

#                 if shadow_info:
#                     # Draw shadow first
#                     draw.text(
#                         (int(pos['x']) + shadow_info['offset_x'], int(pos['y']) + shadow_info['offset_y']), 
#                         str(text_value), 
#                         fill=shadow_info['color'], 
#                         font=styled_font
#                     )

#                 # Draw main text on top
#                 draw.text(
#                     (int(pos['x']), int(pos['y'])), 
#                     str(text_value), 
#                     fill=color, 
#                     font=styled_font
#                 )

#             elif text_value:
#                 # Fallback for unmapped fields
#                 print(f"Warning: No position found for field '{field_name}' with value '{text_value}'")
                
#                 # Default positions for unmapped fields
#                 default_positions = {
#                     'user_message': {'x': 100, 'y': 400, 'fontSize': 30, 'color': 'green'},
#                 }
#                 if field_name in default_positions:
#                     pos = default_positions[field_name]
#                     try:
#                         fallback_font = ImageFont.truetype("arial.ttf", int(pos['fontSize']))
#                     except:
#                         fallback_font = ImageFont.load_default()
                        
#                     draw.text(
#                         (int(pos['x']), int(pos['y'])), 
#                         str(text_value), 
#                         fill=pos['color'],
#                         font=fallback_font
#                     )

#         # Handle doctor image overlay if enabled - ENHANCED
#         print(f"🔍 ===== DOCTOR IMAGE DEBUG =====")
#         print(f"🔍 Template text_positions: {template.text_positions}")
#         print(f"🔍 Content data: {content_data}")

#         image_settings = None

#         # Priority 1: Check content_data for image settings (from live preview)
#         if content_data and 'imageSettings' in content_data:
#             image_settings = content_data['imageSettings']
#             print(f"🔍 Using image settings from content_data: {image_settings}")

#         # Priority 2: Check template for stored image settings
#         elif template.text_positions and 'imageSettings' in template.text_positions:
#             image_settings = template.text_positions['imageSettings']
#             print(f"🔍 Using image settings from template: {image_settings}")

#         print(f"🔍 Final image settings: {image_settings}")

#         # DEBUG: Check doctor image status
#         print(f"🔍 Doctor object: {doctor}")
#         print(f"🔍 Doctor ID: {doctor.id}")
#         print(f"🔍 Doctor name: {doctor.name}")
#         print(f"🔍 Doctor.image field: {doctor.image}")
#         print(f"🔍 Doctor.image exists: {bool(doctor.image)}")
#         print(f"🔍 Doctor.image type: {type(doctor.image)}")

#         if doctor.image:
#             print(f"🔍 Doctor.image.name: {doctor.image.name}")
#             print(f"🔍 Doctor.image.path: {doctor.image.path}")
#             try:
#                 print(f"🔍 Image file exists on disk: {os.path.exists(doctor.image.path)}")
#             except Exception as e:
#                 print(f"🔍 Error checking image path: {e}")
#         else:
#             print(f"🔍 ❌ Doctor has no image file - will use placeholder!")

#         # SINGLE IMAGE PROCESSING SECTION
#         if image_settings and image_settings.get('enabled', False):
#             print(f"🔍 Image settings enabled: {image_settings}")
            
#             try:
#                 # Get positioning settings
#                 img_x = int(image_settings.get('x', 400))
#                 img_y = int(image_settings.get('y', 50))
#                 img_width = int(image_settings.get('width', 150))
#                 img_height = int(image_settings.get('height', 150))
#                 img_fit = image_settings.get('fit', 'cover')
#                 border_radius = int(image_settings.get('borderRadius', 0))
#                 opacity = int(image_settings.get('opacity', 100))
                
#                 # Load or create doctor image
#                 if doctor.image and os.path.exists(doctor.image.path):
#                     # Use actual doctor image
#                     print(f"🔍 Loading actual doctor image from: {doctor.image.path}")
#                     doctor_img = Image.open(doctor.image.path)
                    
#                     # Resize doctor image based on fit mode
#                     if img_fit == 'cover':
#                         doctor_img = doctor_img.resize((img_width, img_height), Image.Resampling.LANCZOS)
#                     elif img_fit == 'contain':
#                         doctor_img.thumbnail((img_width, img_height), Image.Resampling.LANCZOS)
#                     elif img_fit == 'stretch':
#                         doctor_img = doctor_img.resize((img_width, img_height), Image.Resampling.LANCZOS)
                    
#                     print(f"🔍 Loaded and resized actual doctor image")
#                 else:
#                     # Create placeholder image
#                     print(f"🔍 Creating placeholder image {img_width}x{img_height}")
#                     doctor_img = Image.new('RGB', (img_width, img_height), color='#4A90E2')
#                     draw_placeholder = ImageDraw.Draw(doctor_img)
                    
#                     # Add doctor icon/text
#                     try:
#                         # Try to add simple text
#                         font_size = max(20, min(img_width, img_height) // 4)
#                         try:
#                             placeholder_font = ImageFont.truetype("arial.ttf", font_size)
#                         except:
#                             placeholder_font = ImageFont.load_default()
                        
#                         draw_placeholder.text(
#                             (img_width//2, img_height//2), 
#                             "DR", 
#                             fill='white', 
#                             font=placeholder_font,
#                             anchor="mm"
#                         )
#                     except Exception as e:
#                         print(f"🔍 Error adding text to placeholder: {e}")
                    
#                     print(f"🔍 Created placeholder image")
                
#                 # Apply border radius if specified
#                 if border_radius > 0:
#                     mask = Image.new('L', (doctor_img.width, doctor_img.height), 0)
#                     mask_draw = ImageDraw.Draw(mask)
#                     actual_radius = min(doctor_img.width, doctor_img.height) * border_radius // 200
#                     mask_draw.rounded_rectangle(
#                         [(0, 0), (doctor_img.width, doctor_img.height)],
#                         radius=actual_radius,
#                         fill=255
#                     )
                    
#                     # Convert to RGBA and apply mask
#                     if doctor_img.mode != 'RGBA':
#                         doctor_img = doctor_img.convert('RGBA')
#                     doctor_img.putalpha(mask)
#                     print(f"🔍 Applied border radius: {border_radius}%")
                
#                 # Apply opacity
#                 if opacity < 100:
#                     if doctor_img.mode != 'RGBA':
#                         doctor_img = doctor_img.convert('RGBA')
#                     alpha = doctor_img.split()[-1]
#                     alpha = alpha.point(lambda p: int(p * opacity / 100))
#                     doctor_img.putalpha(alpha)
#                     print(f"🔍 Applied opacity: {opacity}%")
                
#                 # Ensure template image is in RGBA mode for compositing
#                 if template_image.mode != 'RGBA':
#                     template_image = template_image.convert('RGBA')
                
#                 # Paste doctor image onto template
#                 if doctor_img.mode == 'RGBA':
#                     template_image.paste(doctor_img, (img_x, img_y), doctor_img)
#                 else:
#                     template_image.paste(doctor_img, (img_x, img_y))
                
#                 print(f"✅ Doctor image composited at ({img_x}, {img_y}) with size {img_width}x{img_height}")
                
#             except Exception as e:
#                 print(f"❌ Error compositing doctor image: {e}")
#                 import traceback
#                 traceback.print_exc()
#         else:
#             print(f"🔍 Image settings not enabled or not found")

#         # Save to temp location
#         output_dir = os.path.join(settings.MEDIA_ROOT, "temp")
#         os.makedirs(output_dir, exist_ok=True)
#         output_path = os.path.join(output_dir, f"temp_image_{uuid.uuid4().hex}.png")
#         template_image.save(output_path)
        
#         return output_path


class GenerateImageContentView(APIView):
    """Generate image with text overlay - DoctorVideo only"""
    
    def post(self, request):
        print(f"🔍 DJANGO: Received POST data: {request.data}")
        
        template_id = request.data.get("template_id")
        doctor_id = request.data.get("doctor_id")  # Existing DoctorVideo ID
        mobile = request.data.get("mobile")        # For creating/finding by mobile
        name = request.data.get("name")            # Required when using mobile
        content_data = request.data.get("content_data", {})
        
        if not template_id:
            return Response({"error": "template_id is required."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Must have either doctor_id OR (mobile + name)
        if not doctor_id and not (mobile and name):
            return Response({"error": "Either doctor_id OR (mobile + name) is required."}, 
                        status=status.HTTP_400_BAD_REQUEST)
        
        try:
            template = VideoTemplates.objects.get(id=template_id, template_type='image')
        except VideoTemplates.DoesNotExist:
            return Response({"error": "Image template not found."}, status=status.HTTP_404_NOT_FOUND)
        
        # SCENARIO 1: Use existing DoctorVideo by ID
        if doctor_id:
            try:
                doctor_video = DoctorVideo.objects.get(id=doctor_id)
                is_new_doctor = False
                print(f"🔍 Using existing DoctorVideo ID: {doctor_id}")
            except DoctorVideo.DoesNotExist:
                return Response({"error": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)
        
        # SCENARIO 2: Find existing or create new DoctorVideo by mobile
        else:
            print(f"🔍 Searching for DoctorVideo with mobile: {mobile}")
            
            # Try to find existing DoctorVideo by mobile
            doctor_video = DoctorVideo.objects.filter(mobile_number=mobile).first()
            
            if doctor_video:
                print(f"🔍 Found existing DoctorVideo: {doctor_video.name} (ID: {doctor_video.id})")
                is_new_doctor = False
            else:
                print(f"🔍 Creating new DoctorVideo for mobile: {mobile}")
                
                # Get employee (required for DoctorVideo)
                try:
                    employee_id = request.data.get("employee_id")
                    if employee_id:
                        employee = Employee.objects.get(employee_id=employee_id)
                    else:
                        employee = Employee.objects.first()
                        if not employee:
                            return Response({"error": "No employees found in system."}, 
                                        status=status.HTTP_404_NOT_FOUND)
                except Employee.DoesNotExist:
                    return Response({"error": f"Employee {employee_id} not found."}, 
                                status=status.HTTP_404_NOT_FOUND)
                
                # Create new DoctorVideo with provided data
                # Get data from content_data (this has the actual form data)
                clinic = content_data.get("doctor_clinic") or request.data.get("clinic", "Unknown Clinic")
                city = content_data.get("doctor_city") or request.data.get("city", "Unknown City") 
                specialization = content_data.get("doctor_specialization") or request.data.get("specialization", "General Medicine")
                state = content_data.get("doctor_state") or request.data.get("state", "Unknown State")

                print(f"🔍 Creating doctor with:")
                print(f"🔍 - Name: {name}")
                print(f"🔍 - Clinic: {clinic}")
                print(f"🔍 - City: {city}")
                print(f"🔍 - Specialization: {specialization}")
                print(f"🔍 - State: {state}")

                # Create new DoctorVideo with proper data
# Create new DoctorVideo with proper data
                doctor_video = DoctorVideo.objects.create(
                    name=name,
                    clinic=clinic,
                    city=city,
                    specialization=specialization,
                    specialization_key=specialization,
                    state=state,
                    mobile_number=mobile,
                    whatsapp_number=mobile,
                    description=request.data.get("description", ""),
                    designation=request.data.get("designation", ""),
                    employee=employee
                )

                print(f"🔍 Created new DoctorVideo: {doctor_video.name} (ID: {doctor_video.id})")

                # HANDLE UPLOADED IMAGE
                uploaded_image = request.FILES.get('doctor_image')
                if uploaded_image:
                    print(f"🔍 ✅ Doctor uploaded image: {uploaded_image.name}")
                    doctor_video.image = uploaded_image
                    doctor_video.save()
                    print(f"🔍 ✅ Image saved to doctor record!")
                else:
                    print(f"🔍 ❌ No image uploaded by doctor")
                    
                    # FALLBACK: Use existing doctor image for testing
                    print(f"🔍 Using fallback image from existing doctor...")
                    source_doctor = DoctorVideo.objects.filter(id=4).first()  # Prathameshhhhh
                    if source_doctor and source_doctor.image:
                        doctor_video.image = source_doctor.image
                        doctor_video.save()
                        print(f"🔍 ⚠️ Using fallback image from {source_doctor.name}")

                print(f"🔍 Final doctor has image: {bool(doctor_video.image)}")
                if doctor_video.image:
                    print(f"🔍 Image path: {doctor_video.image.name}")
                # SMART IMAGE TRANSFER - Choose specific doctor image
                is_new_doctor = True

        if not template.template_image:
                    return Response({"error": "Template does not have an image."}, 
                                status=status.HTTP_400_BAD_REQUEST)
        try:
            # Generate the image
            output_path = self.generate_image_with_text(template, content_data, doctor_video)
            
            # Create ImageContent record
            image_content = ImageContent.objects.create(
                template=template,
                doctor=doctor_video,
                content_data=content_data
            )
            
            # Save generated image
            with open(output_path, 'rb') as f:
                image_content.output_image.save(
                    f"generated_{image_content.id}.png", 
                    File(f), 
                    save=True
                )
            
            os.remove(output_path)
            
            serializer = ImageContentSerializer(image_content, context={'request': request})
            response_data = serializer.data
            response_data['doctor_info'] = {
                'id': doctor_video.id,
                'name': doctor_video.name,
                'clinic': doctor_video.clinic,
                'mobile': doctor_video.mobile_number,
                'has_image': bool(doctor_video.image),
                'is_new_doctor': is_new_doctor
            }
            
            return Response(response_data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"Image generation failed: {e}")
            return Response({"error": "Image generation failed.", "details": str(e)},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            

             
    def generate_image_with_text(self, template, content_data, doctor):
        """Generate image with text overlay using PIL"""
        
        # Load template image
        template_image = Image.open(template.template_image.path)
        draw = ImageDraw.Draw(template_image)
        
        # Get text positions from template
        positions = template.text_positions or {}
        print(f"Template positions: {positions}")
        print(f"Template custom text: {template.custom_text}")
        
        # Font mapping for PIL
        FONT_MAP = {
            'Arial': 'arial.ttf',
            'Times New Roman': 'times.ttf', 
            'Helvetica': 'arial.ttf',
            'Georgia': 'georgia.ttf',
            'Verdana': 'verdana.ttf',
            'Impact': 'impact.ttf',
            'Comic Sans MS': 'comic.ttf'
        }

        def get_font(font_family, font_size):
            """Get PIL font with fallback"""
            font_file = FONT_MAP.get(font_family, 'arial.ttf')
            try:
                return ImageFont.truetype(font_file, font_size)
            except:
                try:
                    return ImageFont.truetype("arial.ttf", font_size)
                except:
                    return ImageFont.load_default()
        
        # Prepare all text data
        all_text_data = {
            'name': content_data.get('doctor_name', doctor.name),
            'clinic': content_data.get('doctor_clinic', doctor.clinic),
            'city': content_data.get('doctor_city', doctor.city),
            'specialization': content_data.get('doctor_specialization', doctor.specialization),
            'state': content_data.get('doctor_state', doctor.state),
            'mobile': doctor.mobile_number,
            'customText': template.custom_text or '', 
        }
        
        print(f"🔍 All text data: {all_text_data}")
                
        # Add text for each field with proper styling
        for field_name, text_value in all_text_data.items():
            if field_name in positions and text_value:
                pos = positions[field_name]
                
                font_size = int(pos.get('fontSize', 40))
                color = pos.get('color', 'black')
                font_weight = pos.get('fontWeight', 'normal')
                font_family = pos.get('fontFamily', 'Arial')

                styled_font = get_font(font_family, font_size)

                if font_weight == 'bold':
                    try:
                        bold_font_file = FONT_MAP.get(font_family, 'arial.ttf').replace('.ttf', 'bd.ttf')
                        styled_font = ImageFont.truetype(bold_font_file, font_size)
                    except:
                        styled_font = get_font(font_family, font_size)

                # Handle text shadow
                text_shadow = pos.get('textShadow', 'none')
                if text_shadow != 'none':
                    shadow_info = parse_css_shadow(text_shadow)
                    if shadow_info:
                        draw.text(
                            (int(pos['x']) + shadow_info['offset_x'], int(pos['y']) + shadow_info['offset_y']), 
                            str(text_value), 
                            fill=shadow_info['color'], 
                            font=styled_font
                        )

                # Draw main text
                draw.text(
                    (int(pos['x']), int(pos['y'])), 
                    str(text_value), 
                    fill=color, 
                    font=styled_font
                )

            elif text_value:
                print(f"Warning: No position found for field '{field_name}' with value '{text_value}'")

        # Handle doctor image overlay
        print(f"🔍 ===== DOCTOR IMAGE DEBUG =====")
        print(f"🔍 Doctor ID: {doctor.id}")
        print(f"🔍 Doctor name: {doctor.name}")
        print(f"🔍 Doctor.image field: {doctor.image}")
        print(f"🔍 Doctor.image exists: {bool(doctor.image)}")

        # ADD DEBUG FOR ALL DOCTORS WITH IMAGES
        # print(f"🔍 === CHECKING ALL DOCTORS WITH IMAGES ===")
        # all_doctors_with_images = DoctorVideo.objects.exclude(image='').exclude(image__isnull=True)
        # print(f"🔍 Total doctors with images: {all_doctors_with_images.count()}")
        
        # for doc in all_doctors_with_images[:5]:  # Show first 5
        #     print(f"🔍 Doctor {doc.id}: {doc.name} - Image: {doc.image.name}")
        #     if doc.image:
        #         try:
        #             print(f"🔍 - Image path exists: {os.path.exists(doc.image.path)}")
        #         except:
        #             print(f"🔍 - Error checking image path")

        # Get image settings
        image_settings = None
        if content_data and 'imageSettings' in content_data:
            image_settings = content_data['imageSettings']
            print(f"🔍 Using image settings from content_data: {image_settings}")
        elif template.text_positions and 'imageSettings' in template.text_positions:
            image_settings = template.text_positions['imageSettings']
            print(f"🔍 Using image settings from template: {image_settings}")

        # Process doctor image if settings enabled
        if image_settings and image_settings.get('enabled', False):
            print(f"🔍 Image settings enabled: {image_settings}")
            
            try:
                img_x = int(image_settings.get('x', 400))
                img_y = int(image_settings.get('y', 50))
                img_width = int(image_settings.get('width', 150))
                img_height = int(image_settings.get('height', 150))
                img_fit = image_settings.get('fit', 'cover')
                border_radius = int(image_settings.get('borderRadius', 0))
                opacity = int(image_settings.get('opacity', 100))
                
                # Check if doctor has actual image
                if doctor.image and os.path.exists(doctor.image.path):
                    print(f"🔍 Loading actual doctor image from: {doctor.image.path}")
                    doctor_img = Image.open(doctor.image.path)
                    
                    # Resize based on fit mode
                    if img_fit == 'cover':
                        doctor_img = doctor_img.resize((img_width, img_height), Image.Resampling.LANCZOS)
                    elif img_fit == 'contain':
                        doctor_img.thumbnail((img_width, img_height), Image.Resampling.LANCZOS)
                    elif img_fit == 'stretch':
                        doctor_img = doctor_img.resize((img_width, img_height), Image.Resampling.LANCZOS)
                    
                    print(f"🔍 ✅ Using REAL doctor image")
                else:
                    # Create placeholder
                    print(f"🔍 Creating placeholder image {img_width}x{img_height}")
                    doctor_img = Image.new('RGB', (img_width, img_height), color='#4A90E2')
                    draw_placeholder = ImageDraw.Draw(doctor_img)
                    
                    font_size = max(20, min(img_width, img_height) // 4)
                    try:
                        placeholder_font = ImageFont.truetype("arial.ttf", font_size)
                    except:
                        placeholder_font = ImageFont.load_default()
                    
                    draw_placeholder.text(
                        (img_width//2, img_height//2), 
                        "DR", 
                        fill='white', 
                        font=placeholder_font,
                        anchor="mm"
                    )
                    print(f"🔍 ⚠️ Using PLACEHOLDER image")
                
                # Apply styling
                if border_radius > 0:
                    mask = Image.new('L', (doctor_img.width, doctor_img.height), 0)
                    mask_draw = ImageDraw.Draw(mask)
                    actual_radius = min(doctor_img.width, doctor_img.height) * border_radius // 200
                    mask_draw.rounded_rectangle(
                        [(0, 0), (doctor_img.width, doctor_img.height)],
                        radius=actual_radius,
                        fill=255
                    )
                    
                    if doctor_img.mode != 'RGBA':
                        doctor_img = doctor_img.convert('RGBA')
                    doctor_img.putalpha(mask)
                
                if opacity < 100:
                    if doctor_img.mode != 'RGBA':
                        doctor_img = doctor_img.convert('RGBA')
                    alpha = doctor_img.split()[-1]
                    alpha = alpha.point(lambda p: int(p * opacity / 100))
                    doctor_img.putalpha(alpha)
                
                # Composite onto template
                if template_image.mode != 'RGBA':
                    template_image = template_image.convert('RGBA')
                
                if doctor_img.mode == 'RGBA':
                    template_image.paste(doctor_img, (img_x, img_y), doctor_img)
                else:
                    template_image.paste(doctor_img, (img_x, img_y))
                
                print(f"✅ Doctor image composited at ({img_x}, {img_y}) with size {img_width}x{img_height}")
                
            except Exception as e:
                print(f"❌ Error compositing doctor image: {e}")
                import traceback
                traceback.print_exc()

        # Save output
        output_dir = os.path.join(settings.MEDIA_ROOT, "temp")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"temp_image_{uuid.uuid4().hex}.png")
        template_image.save(output_path)
        
        return output_path

class ImageContentListView(APIView):
    """List generated image contents with pagination"""
    
    def get(self, request):
        doctor_id = request.GET.get('doctor_id')  # CHANGE: doctor_id instead of employee_id
        template_id = request.GET.get('template_id')
        
        contents = ImageContent.objects.all()
        
        if doctor_id:
            try:
                doctor = DoctorVideo.objects.get(id=doctor_id)
                contents = contents.filter(doctor=doctor)
            except DoctorVideo.DoesNotExist:
                return Response({"error": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)
        
        if template_id:
            contents = contents.filter(template_id=template_id)
        
        # Apply pagination
        paginator = Pagination_class()
        page = paginator.paginate_queryset(contents, request)
        
        serializer = ImageContentSerializer(page, many=True, context={'request': request})
        return paginator.get_paginated_response(serializer.data)
    
class DoctorSearchView(APIView):
    """Search doctor by mobile number or name"""
    
    def get(self, request):
        mobile = request.GET.get('mobile')
        
        if not mobile:
            return Response({"error": "mobile parameter required."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Search in DoctorVideo first
        doctor_video = DoctorVideo.objects.filter(mobile_number=mobile).first()
        
        if doctor_video:
            return Response({
                "found": True,
                "doctor": {
                    "id": doctor_video.id,
                    "name": doctor_video.name,
                    "clinic": doctor_video.clinic,
                    "city": doctor_video.city,
                    "mobile": doctor_video.mobile_number,
                    "specialization": doctor_video.specialization,
                    "state": doctor_video.state,
                    "model": "DoctorVideo"
                }
            })
        
        # If not found in DoctorVideo, search in Doctor model
        doctor = Doctor.objects.filter(mobile_number=mobile).first()
        
        if doctor:
            return Response({
                "found": True,
                "doctor": {
                    "id": doctor.id,
                    "name": doctor.name,
                    "clinic": doctor.clinic,
                    "city": doctor.city,
                    "mobile": doctor.mobile_number,
                    "specialization": doctor.specialization,
                    "model": "Doctor"
                },
                "note": "Doctor found in main database. Will be converted to DoctorVideo when image is generated."
            })
        
        return Response({"found": False})



@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def AddEmployeeTemplates(request, template_type='video'):
    """Handle both video and image template creation"""
    try:
        data = request.data.copy()
        data['template_type'] = template_type
        
        # Handle image template positioning data
        if template_type == 'image':
            # Combine text_positions and imageSettings into one JSON field
            text_positions = data.get('text_positions', '{}')
            image_settings = data.get('imageSettings', '{}')
            
            # Parse JSON strings if they come as strings
            if isinstance(text_positions, str):
                import json
                text_positions = json.loads(text_positions)
            if isinstance(image_settings, str):
                import json
                image_settings = json.loads(image_settings)
            
            # Combine both positioning data
            combined_positions = text_positions.copy() if text_positions else {}
            combined_positions['imageSettings'] = image_settings
            
            # Update data with combined positions
            data['text_positions'] = json.dumps(combined_positions)
            
            print(f"🔍 Combined positioning data: {combined_positions}")
            
        print(f"🔍 Received template_type: {template_type}")
        print(f"🔍 Received custom_text: {data.get('custom_text', 'NOT_FOUND')}")

        if template_type == 'image':
            serializer = ImageTemplateSerializer(data=data, context={'request': request})
        else:
            serializer = VideoTemplatesSerializer(data=data)
            
        if serializer.is_valid():
            template = serializer.save()
            print(f"🔍 Template saved with ID: {template.id}")
            print(f"🔍 Saved custom_text: '{template.custom_text}'")
            print(f"🔍 Saved text_positions: {template.text_positions}")
            
            return Response({
                'id': template.id,
                'message': f'{template_type.title()} template created successfully',
                'status': template.status
            }, status=status.HTTP_201_CREATED)
        else:
            return Response({
                'error': 'Validation failed',
                'details': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        return Response({
            'error': 'Template creation failed',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['GET'])
def getFilteredVideoTemplates(request):
    """Get filtered video templates by status"""
    status_param = request.GET.get('status', 'true')
    template_type = request.GET.get('template_type', 'video')
    
    try:
        status_bool = bool(strtobool(status_param))
        templates = VideoTemplates.objects.filter(
            status=status_bool,
            template_type=template_type
        ).order_by('-created_at')
        
        if template_type == 'image':
            serializer = ImageTemplateSerializer(templates, many=True, context={'request': request})
        else:
            serializer = VideoTemplatesSerializer(templates, many=True)
            
        return Response(serializer.data, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


#!DONE -04

#! AFTER COMPARIOSN

class DoctorUpdateDeleteView(APIView):
    # permission_classes = [IsAuthenticated]
    permission_classes = []  # EXPLICITLY OVERRIDE GLOBAL AUTH
    authentication_classes = []
    
    def get_doctor(self, doctor_id, employee_id):
        """Get doctor with permission check"""
        try:
            doctor = DoctorVideo.objects.get(id=doctor_id)
            
            # Get current employee
            try:
                current_employee = Employee.objects.get(employee_id=employee_id)
            except Employee.DoesNotExist:
                raise PermissionError("Employee not found")
            
            # Admin can edit any doctor, Employee can only edit their own
            if current_employee.user_type != 'Admin' and doctor.employee != current_employee:
                raise PermissionError("You can only edit your own doctors")
            
            return doctor
        except DoctorVideo.DoesNotExist:
            raise Http404("Doctor not found")
    
    def patch(self, request, doctor_id):
        """Update doctor profile"""
        try:
            employee_id = request.data.get('employee_id')
            if not employee_id:
                return Response({'error': 'employee_id is required'}, status=status.HTTP_400_BAD_REQUEST)
                
            doctor = self.get_doctor(doctor_id, employee_id)
            
            # Update fields
            serializer = DoctorVideoSerializer(doctor, data=request.data, partial=True)
            if serializer.is_valid():
                updated_doctor = serializer.save()
                
                return Response({
                    'status': 'success',
                    'message': 'Doctor updated successfully',
                    'data': DoctorVideoSerializer(updated_doctor).data
                })
            else:
                return Response({
                    'status': 'error',
                    'errors': serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except PermissionError as e:
            return Response({'error': str(e)}, status=status.HTTP_403_FORBIDDEN)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request, doctor_id):
        """Delete doctor and all associated content"""
        try:
            employee_id = request.query_params.get('employee_id')  # For DELETE, use query params
            if not employee_id:
                return Response({'error': 'employee_id is required'}, status=status.HTTP_400_BAD_REQUEST)
                
            doctor = self.get_doctor(doctor_id, employee_id)
            doctor_name = doctor.name
            
            # Delete associated content first
            DoctorOutputVideo.objects.filter(doctor=doctor).delete()
            ImageContent.objects.filter(doctor=doctor).delete()
            
            # Delete doctor
            doctor.delete()
            
            return Response({
                'status': 'success',
                'message': f'Doctor {doctor_name} and all associated content deleted successfully'
            })
            
        except PermissionError as e:
            return Response({'error': str(e)}, status=status.HTTP_403_FORBIDDEN)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class RegenerateContentView(APIView):
    # permission_classes = [IsAuthenticated]
    permission_classes = []  # EXPLICITLY OVERRIDE GLOBAL AUTH
    authentication_classes = []

    def post(self, request):
        print("=== [RegenerateContentView] POST called ===")
        print("Request data:", dict(request.data))
        """Regenerate video/image with different template"""
        doctor_id = request.data.get('doctor_id')
        template_id = request.data.get('template_id')
        content_type = request.data.get('content_type')  # 'video' or 'image'
        
        if not all([doctor_id, template_id, content_type]):
            print("Missing required fields: doctor_id, template_id, content_type")
            return Response({
                'error': 'doctor_id, template_id, and content_type are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            print(f"Looking for DoctorVideo with id={doctor_id}")
            doctor = DoctorVideo.objects.get(id=doctor_id)
            
            # Permission check
            employee_id = request.data.get('employee_id')
            print("Employee ID received:", employee_id)
            if not employee_id:
                print("employee_id is required but not provided!")
                return Response({'error': 'employee_id is required'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                print(f"Looking for Employee with employee_id={employee_id}")
                current_employee = Employee.objects.get(employee_id=employee_id)
            except Employee.DoesNotExist:
                print(f"Employee not found: {employee_id}")
                return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

            if current_employee.user_type != 'Admin' and doctor.employee != current_employee:
                print(f"Permission denied: employee {employee_id} is not admin and not owner")
                return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)
            
            print(f"Looking for VideoTemplates with id={template_id}")
            template = VideoTemplates.objects.get(id=template_id)
            
            print(f"Requested content_type: {content_type}")
            if content_type == 'video':
                print("[REGENERATE] Video branch")
                # Use existing video generation logic
                return self.regenerate_video(doctor, template)
            elif content_type == 'image':
                print("[REGENERATE] Image branch")
                # Use existing image generation logic
                return self.regenerate_image(doctor, template, request.data.get('content_data', {}))
            else:
                print("Invalid content_type provided:", content_type)
                return Response({'error': 'Invalid content_type'}, status=status.HTTP_400_BAD_REQUEST)
                
        except DoctorVideo.DoesNotExist:
            print(f"Doctor not found: {doctor_id}")
            return Response({'error': 'Doctor not found'}, status=status.HTTP_404_NOT_FOUND)
        except VideoTemplates.DoesNotExist:
            print(f"Template not found: {template_id}")
            return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print("EXCEPTION in RegenerateContentView:", str(e))
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
#Byvasi
    def regenerate_video(self, doctor, template):
        print("=== [regenerate_video] called ===")
        print("doctor:", doctor)
        print("template:", template)
        if not doctor.image:
            print("Doctor has no image for video generation!")
            return Response({'error': 'Doctor has no image for video generation'}, 
                        status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Use existing video generation logic
            output_filename = f"{doctor.id}_{template.id}_{uuid.uuid4().hex[:8]}_output.mp4"
            output_dir = os.path.join(settings.MEDIA_ROOT, "output", str(doctor.employee.id), str(doctor.id))
            print("Creating output directory if not exists:", output_dir)
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, output_filename)
            print("Output path for generated video:", output_path)
            
            # Call video generation (reuse existing method)
            print("Calling generate_custom_video...")
            video_gen = VideoGenViewSet()
            video_gen.generate_custom_video(
                main_video_path=template.template_video.path,
                image_path=doctor.image.path,
                name=doctor.name,
                clinic=doctor.clinic,
                city=doctor.city,
                specialization_key=doctor.specialization_key,
                state=doctor.state,
                output_path=output_path
            )
            print("Video generated successfully, now saving DoctorOutputVideo...")
            # Save new video
            relative_path = os.path.relpath(output_path, settings.MEDIA_ROOT)
            print("Relative output path:", relative_path)
            output_video = DoctorOutputVideo.objects.create(
                doctor=doctor,
                template=template,
                video_file=relative_path
            )
            print("DoctorOutputVideo created:", output_video.id)
            
            return Response({
                'status': 'success',
                'message': 'Video regenerated successfully',
                'data': DoctorOutputVideoSerializer(output_video).data
            })
            
        except Exception as e:
            print("Exception in regenerate_video:", str(e))
            import traceback
            traceback.print_exc()
            return Response({'error': f'Video generation failed: {str(e)}'}, 
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def regenerate_image(self, doctor, template, content_data):
        print("=== [regenerate_image] called ===")
        print("doctor:", doctor)
        print("template:", template)
        print("content_data:", content_data)
        try:
            # Use existing image generation logic
            print("Calling generate_image_with_text...")
            image_gen = GenerateImageContentView()
            output_path = image_gen.generate_image_with_text(template, content_data, doctor)
            print("Image generated at:", output_path)
            
            # Create new image content
            print("Creating ImageContent object...")
            image_content = ImageContent.objects.create(
                template=template,
                doctor=doctor,
                content_data=content_data
            )
            
            # Save generated image
            print("Saving generated image to ImageContent...")
            with open(output_path, 'rb') as f:
                image_content.output_image.save(
                    f"regenerated_{image_content.id}.png",
                    File(f),
                    save=True
                )
            
            print("Removing temporary output image file...")
            os.remove(output_path)
            
            print("Image regenerated and saved successfully!")
            return Response({
                'status': 'success',
                'message': 'Image regenerated successfully',
                'data': ImageContentSerializer(image_content).data
            })
            
        except Exception as e:
            print("Exception in regenerate_image:", str(e))
            import traceback
            traceback.print_exc()
            return Response({'error': f'Image generation failed: {str(e)}'}, 
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)


    
    # def regenerate_video(self, doctor, template):
    #     """Regenerate video with new template"""
    #     if not doctor.image:
    #         return Response({'error': 'Doctor has no image for video generation'}, 
    #                       status=status.HTTP_400_BAD_REQUEST)
        
    #     try:
    #         # Use existing video generation logic
    #         output_filename = f"{doctor.id}_{template.id}_{uuid.uuid4().hex[:8]}_output.mp4"
    #         output_dir = os.path.join(settings.MEDIA_ROOT, "output", str(doctor.employee.id), str(doctor.id))
    #         os.makedirs(output_dir, exist_ok=True)
    #         output_path = os.path.join(output_dir, output_filename)
            
    #         # Call video generation (reuse existing method)
    #         video_gen = VideoGenViewSet()
    #         video_gen.generate_custom_video(
    #             main_video_path=template.template_video.path,
    #             image_path=doctor.image.path,
    #             name=doctor.name,
    #             clinic=doctor.clinic,
    #             city=doctor.city,
    #             specialization_key=doctor.specialization_key,
    #             state=doctor.state,
    #             output_path=output_path
    #         )
            
    #         # Save new video
    #         relative_path = os.path.relpath(output_path, settings.MEDIA_ROOT)
    #         output_video = DoctorOutputVideo.objects.create(
    #             doctor=doctor,
    #             template=template,
    #             video_file=relative_path
    #         )
            
    #         return Response({
    #             'status': 'success',
    #             'message': 'Video regenerated successfully',
    #             'data': DoctorOutputVideoSerializer(output_video).data
    #         })
            
    #     except Exception as e:
    #         return Response({'error': f'Video generation failed: {str(e)}'}, 
    #                       status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    # def regenerate_image(self, doctor, template, content_data):
    #     """Regenerate image with new template"""
    #     try:
    #         # Use existing image generation logic
    #         image_gen = GenerateImageContentView()
    #         output_path = image_gen.generate_image_with_text(template, content_data, doctor)
            
    #         # Create new image content
    #         image_content = ImageContent.objects.create(
    #             template=template,
    #             doctor=doctor,
    #             content_data=content_data
    #         )
            
    #         # Save generated image
    #         with open(output_path, 'rb') as f:
    #             image_content.output_image.save(
    #                 f"regenerated_{image_content.id}.png",
    #                 File(f),
    #                 save=True
    #             )
            
    #         os.remove(output_path)
            
    #         return Response({
    #             'status': 'success',
    #             'message': 'Image regenerated successfully',
    #             'data': ImageContentSerializer(image_content).data
    #         })
            
    #     except Exception as e:
    #         return Response({'error': f'Image generation failed: {str(e)}'}, 
    #                       status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class DeleteContentView(APIView):
    # permission_classes = [IsAuthenticated]
    permission_classes = []  # EXPLICITLY OVERRIDE GLOBAL AUTH
    authentication_classes = []  
    def delete(self, request, content_type, content_id):
        """Delete specific video or image content"""
        try:
            if content_type == 'video':
                content = DoctorOutputVideo.objects.get(id=content_id)
                doctor = content.doctor
            elif content_type == 'image':
                content = ImageContent.objects.get(id=content_id)
                doctor = content.doctor
            else:
                return Response({'error': 'Invalid content type'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Permission check
            # Get employee_id from query params for DELETE requests
            employee_id = request.query_params.get('employee_id')
            if not employee_id:
                return Response({'error': 'employee_id is required'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                current_employee = Employee.objects.get(employee_id=employee_id)
            except Employee.DoesNotExist:
                return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

            # Permission check using Employee model
            if current_employee.user_type != 'Admin' and doctor.employee != current_employee:
                return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)
            
            # Delete the content
            content.delete()
            
            return Response({
                'status': 'success',
                'message': f'{content_type.title()} deleted successfully'
            })
            
        except (DoctorOutputVideo.DoesNotExist, ImageContent.DoesNotExist):
            return Response({'error': 'Content not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


#! on SUNDAY