import os
import json
import uuid
import random
import string
import logging
import subprocess
from datetime import datetime, date, timedelta
import mimetypes
from django.core.files.storage import default_storage
from django.core.exceptions import ValidationError

import openpyxl
import pandas as pd  # type: ignore
from PIL import Image, ImageDraw, ImageFont

from django.conf import settings
from django.core.files import File
from django.db import IntegrityError
from django.db.models import Count, Q
from django.contrib.postgres.aggregates import ArrayAgg
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone

from rest_framework import status, viewsets, generics
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
#from django_ratelimit.decorators import ratelimit
from django.utils.decorators import method_decorator
#from django_ratelimit.decorators import ratelimit
import psutil
import threading
import time
from django.db import models
from django.core.cache import cache

# Resource monitoring decorator
def monitor_resources(func):
    def wrapper(*args, **kwargs):
        start_memory = psutil.virtual_memory().percent
        start_time = time.time()

        try:
            result = func(*args, **kwargs)
            return result
        finally:
            end_memory = psutil.virtual_memory().percent
            duration = time.time() - start_time

            # Log if resource usage is high
            if end_memory - start_memory > 10 or duration > 30:
                logger.warning(f"High resource usage in {func.__name__}: "
                             f"Memory: {start_memory}% -> {end_memory}%, "
                             f"Duration: {duration:.2f}s")
    return wrapper



from django.core.cache import cache
#from rest_framework.throttling import UserRateThrottle, AnonRateThrottle

#class BurstRateThrottle(UserRateThrottle):
#    scope = 'burst'
#    rate = '30/min'  # Reduced from 60 to prevent spam
#
#class SustainedRateThrottle(UserRateThrottle):
#    scope = 'sustained'
#    rate = '1000/day'  # Keep this - reasonable daily limit
#
#class MediaGenerationThrottle(UserRateThrottle):
#    scope = 'media_generation'
#    rate = '500/hour'  # Increased from 10 to 500 - much more user-friendly
#
from rest_framework.pagination import PageNumberPagination
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenRefreshView
from distutils.util import strtobool
from .models import (
    Employee,
    DoctorVideo,
    EmployeeLoginHistory,
    VideoTemplates,
    ImageContent,
    Brand,
    Designation,
    # DoctorUsageHistory


)

from .serializers import (
    EmployeeLoginSerializer,
    EmployeeSerializer,
    DoctorSerializer,
    DoctorVideoSerializer,
    VideoTemplatesSerializer,
    ImageContentSerializer,
    ImageTemplateSerializer,
    BrandSerializer,

)
import psutil
import os
from django.core.cache import cache
from celery import current_app

class HealthCheckView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        """System health check for monitoring"""
        health_data = {
            "status": "healthy",
            "timestamp": timezone.now().isoformat(),
            "checks": {}
        }

        overall_status = True

        # Database check
        try:
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
                health_data["checks"]["database"] = {"status": "ok"}
        except Exception as e:
            health_data["checks"]["database"] = {"status": "error", "message": str(e)}
            overall_status = False

        # Cache check
        health_data["checks"]["cache"] = {"status": "disabled", "message": "Cache not in use"}

        # Celery check
        try:
            inspector = current_app.control.inspect()
            active_workers = inspector.active()
            if active_workers:
                health_data["checks"]["celery"] = {"status": "ok", "workers": len(active_workers)}
            else:
                health_data["checks"]["celery"] = {"status": "warning", "message": "no active workers"}
        except Exception as e:
            health_data["checks"]["celery"] = {"status": "error", "message": str(e)}

        # Disk space check
        try:
            disk_usage = psutil.disk_usage(settings.MEDIA_ROOT)
            free_percent = (disk_usage.free / disk_usage.total) * 100
            if free_percent < 10:
                health_data["checks"]["disk"] = {"status": "warning", "free_percent": round(free_percent, 2)}
            else:
                health_data["checks"]["disk"] = {"status": "ok", "free_percent": round(free_percent, 2)}
        except Exception as e:
            health_data["checks"]["disk"] = {"status": "error", "message": str(e)}

        # Memory check
        try:
            memory = psutil.virtual_memory()
            if memory.percent > 90:
                health_data["checks"]["memory"] = {"status": "warning", "usage_percent": memory.percent}
                overall_status = False
            else:
                health_data["checks"]["memory"] = {"status": "ok", "usage_percent": memory.percent}
        except Exception as e:
            health_data["checks"]["memory"] = {"status": "error", "message": str(e)}

        health_data["status"] = "healthy" if overall_status else "unhealthy"
        status_code = status.HTTP_200_OK if overall_status else status.HTTP_503_SERVICE_UNAVAILABLE

# Celery task
try:
    from employee_app.tasks import generate_custom_video_task
except Exception:
    class _NoopTask:
        def delay(self, *a, **kw):
            pass
    generate_custom_video_task = _NoopTask()

# ------------------------------------------------------------------------------
# Logging
# ------------------------------------------------------------------------------
logger = logging.getLogger(__name__)


# Security Configuration
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
ALLOWED_IMAGE_TYPES = ['image/jpeg', 'image/png', 'image/jpg']
ALLOWED_VIDEO_TYPES = ['video/mp4', 'video/avi', 'video/mov']


import psutil
import os
from django.core.cache import cache
from celery import current_app

class HealthCheckView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        """System health check for monitoring"""
        health_data = {
            "status": "healthy",
            "timestamp": timezone.now().isoformat(),
            "checks": {}
        }

        overall_status = True

        # Database check
        try:
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
                health_data["checks"]["database"] = {"status": "ok"}
        except Exception as e:
            health_data["checks"]["database"] = {"status": "error", "message": str(e)}
            overall_status = False

        # Cache check
        try:
            cache.set("health_check", "ok", 10)
            if cache.get("health_check") == "ok":
                health_data["checks"]["cache"] = {"status": "ok"}
            else:
                health_data["checks"]["cache"] = {"status": "error", "message": "Cache not responding"}
                overall_status = False
        except Exception as e:
            health_data["checks"]["cache"] = {"status": "error", "message": str(e)}
            overall_status = False

        # Celery check
        try:
            inspector = current_app.control.inspect()
            active_workers = inspector.active()
            if active_workers:
                health_data["checks"]["celery"] = {"status": "ok", "workers": len(active_workers)}
            else:
                health_data["checks"]["celery"] = {"status": "warning", "message": "no active workers"}
        except Exception as e:
            health_data["checks"]["celery"] = {"status": "error", "message": str(e)}

        # Disk space check
        try:
            disk_usage = psutil.disk_usage(settings.MEDIA_ROOT)
            free_percent = (disk_usage.free / disk_usage.total) * 100
            if free_percent < 10:
                health_data["checks"]["disk"] = {"status": "warning", "free_percent": round(free_percent, 2)}
            else:
                health_data["checks"]["disk"] = {"status": "ok", "free_percent": round(free_percent, 2)}
        except Exception as e:
            health_data["checks"]["disk"] = {"status": "error", "message": str(e)}

        # Memory check
        try:
            memory = psutil.virtual_memory()
            if memory.percent > 90:
                health_data["checks"]["memory"] = {"status": "warning", "usage_percent": memory.percent}
                overall_status = False
            else:
                health_data["checks"]["memory"] = {"status": "ok", "usage_percent": memory.percent}
        except Exception as e:
            health_data["checks"]["memory"] = {"status": "error", "message": str(e)}

        health_data["status"] = "healthy" if overall_status else "unhealthy"
        status_code = status.HTTP_200_OK if overall_status else status.HTTP_503_SERVICE_UNAVAILABLE

        return Response(health_data, status=status_code)
@api_view(['GET'])
def system_metrics(request):
    """Detailed system metrics for monitoring"""
    try:
        import psutil

        # Memory metrics
        memory = psutil.virtual_memory()

        # Disk metrics
        disk = psutil.disk_usage(settings.MEDIA_ROOT)

        # CPU metrics
        cpu_percent = psutil.cpu_percent(interval=1)

        # Active connections
        connections = len(psutil.net_connections())

        # Celery metrics
        try:
            from celery import current_app
            inspector = current_app.control.inspect()
            active_tasks = inspector.active()
            active_count = sum(len(tasks) for tasks in (active_tasks or {}).values())
        except:
            active_count = 0

        metrics = {
            "timestamp": timezone.now().isoformat(),
            "memory": {
                "percent": memory.percent,
                "available_gb": round(memory.available / (1024**3), 2),
                "used_gb": round(memory.used / (1024**3), 2)
            },
            "disk": {
                "percent": round((disk.used / disk.total) * 100, 2),
                "free_gb": round(disk.free / (1024**3), 2)
            },
            "cpu_percent": cpu_percent,
            "active_connections": connections,
            "active_celery_tasks": active_count
        }

        return Response(metrics)

    except Exception as e:
        return Response({"error": str(e)}, status=500)




# Add process monitoring
def get_system_stats():
    """Get basic system statistics for monitoring"""
    try:
        import psutil
        return {
            'cpu_percent': psutil.cpu_percent(interval=1),
            'memory_percent': psutil.virtual_memory().percent,
            'disk_usage': psutil.disk_usage('/').percent if psutil.disk_usage('/') else 0,
            'active_connections': len(psutil.net_connections()),
        }
    except Exception:
        return {}


def validate_file_upload(file):
    """Validate uploaded files for security"""
    if not file:
        raise ValidationError("No file provided")

    if file.size > MAX_FILE_SIZE:
        raise ValidationError(f"File too large. Maximum size is {MAX_FILE_SIZE/1024/1024}MB")

    # Additional size check for images
    if file.content_type and file.content_type.startswith('image/'):
        if file.size > 5 * 1024 * 1024:  # 5MB for images
            raise ValidationError("Image files must be under 5MB")

    # Check MIME type
    mime_type, _ = mimetypes.guess_type(file.name)
    if mime_type not in ALLOWED_IMAGE_TYPES + ALLOWED_VIDEO_TYPES:
        raise ValidationError("Invalid file type")

    # Check file extension
    allowed_extensions = ['.jpg', '.jpeg', '.png', '.mp4', '.avi', '.mov']
    if not any(file.name.lower().endswith(ext) for ext in allowed_extensions):
        raise ValidationError("Invalid file extension")

    return True

# ------------------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------------------

def parse_css_shadow(shadow_str):
    """Parse CSS-like text-shadow: '2px 2px 4px rgba(0,0,0,0.7)' -> dict or None"""
    if shadow_str == 'none' or not shadow_str:
        return None
    try:
        if 'rgba' in shadow_str:
            rgba_start = shadow_str.find('rgba(')
            rgba_end = shadow_str.find(')', rgba_start)
            rgba_str = shadow_str[rgba_start+5:rgba_end]
            rgba_values = [float(x.strip()) for x in rgba_str.split(',')]
            shadow_color = (int(rgba_values[0]), int(rgba_values[1]), int(rgba_values[2]))
            offset_part = shadow_str[:rgba_start].strip()
            offsets = offset_part.replace('px', '').split()
            if len(offsets) >= 2:
                return {
                    'offset_x': int(float(offsets[0])),
                    'offset_y': int(float(offsets[1])),
                    'color': shadow_color
                }
        elif 'px' in shadow_str:
            parts = shadow_str.replace('px', '').split()
            if len(parts) >= 2:
                return {
                    'offset_x': int(float(parts[0])),
                    'offset_y': int(float(parts[1])),
                    'color': (128, 128, 128)
                }
    except:  # noqa: E722
        pass
    return {'offset_x': 2, 'offset_y': 2, 'color': (128, 128, 128)}


def _ff_esc(s: str) -> str:
    s = str(s or "")
    s = s.replace("\\", "\\\\")  # backslash first
    s = s.replace(":", r"\:")
    s = s.replace("'", r"\'")
    s = s.replace("%", r"\%")
    return s

def _num_or_expr(v, default="0"):
    v = str(v).strip()
    return v if v and not v.lstrip("+-").isdigit() else str(int(v or default))

# ------------------------------------------------------------------------------
# Pagination
# ------------------------------------------------------------------------------

class Pagination_class(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

# ------------------------------------------------------------------------------
# Employees & Auth
# ------------------------------------------------------------------------------

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer

def get_tokens_for_employee(employee):
    from django.contrib.auth.models import User

    # Create or get a Django User for this employee
    user, created = User.objects.get_or_create(
        username=employee.employee_id,
        defaults={
            'email': employee.email or '',
            'first_name': employee.first_name,
            'last_name': employee.last_name or ''
        }
    )

    refresh = RefreshToken.for_user(user)
    refresh.access_token.set_exp(lifetime=timedelta(hours=1))
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
        'access_token_exp': refresh.access_token.payload['exp'],
    }

@permission_classes([AllowAny])
@api_view(['POST'])
def employee_login_api(request):
    employee_id = request.data.get('employee_id')
    rbm_region = request.data.get('rbm_region')
    
    if not employee_id or not rbm_region:
        return Response({
            'status': 'error',
            'message': 'Both employee_id and rbm_region are required'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Validate designation-RBM combination
        designation = Designation.objects.get(login_code=employee_id, rbm_region=rbm_region)
    except Designation.DoesNotExist:
        return Response({
            'status': 'error',
            'message': f'Invalid combination: {employee_id} does not belong to {rbm_region}'
        }, status=status.HTTP_401_UNAUTHORIZED)

    try:
        # Check if employee already exists
        employee = Employee.objects.get(employee_id=employee_id)
        
        if not employee.status:
            return Response({
                'status': 'error',
                'message': 'Your account is inactive. Please contact the admin department.'
            }, status=status.HTTP_403_FORBIDDEN)

        # Update employee with designation info
        employee.designation = employee_id
        employee.rbm_region = rbm_region
        employee.login_date = timezone.now()
        employee.has_logged_in = True
        employee.save(update_fields=['designation', 'rbm_region', 'login_date', 'has_logged_in'])

    except Employee.DoesNotExist:
        # Create new employee if doesn't exist
        employee = Employee.objects.create(
            employee_id=employee_id,
            first_name=employee_id.split('-')[0],  # Extract prefix as first name
            designation=employee_id,
            rbm_region=rbm_region,
            user_type='Employee',
            has_logged_in=True
        )

    tokens = get_tokens_for_employee(employee)

    EmployeeLoginHistory.objects.create(
        employee=employee,
        employee_identifier=employee.employee_id,
        name=f"{employee.first_name} {employee.last_name or ''}",
        email=employee.email,
        phone=employee.phone,
        department=employee.department,
        user_type=employee.user_type
    )

    return Response({
        'status': 'success',
        'message': 'Login successful',
        'tokens': tokens,
        'employee': {
            'id': employee.id,
            'employee_id': employee.employee_id,
            'name': f"{employee.first_name} {employee.last_name or ''}",
            'email': employee.email,
            'department': employee.department,
            'first_name': employee.first_name,
            'last_name': employee.last_name or '',
            'user_type': employee.user_type,
            'designation': employee.designation,
            'rbm_region': employee.rbm_region
        }
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([AllowAny])
def get_rbm_regions(request):
    """Get all unique RBM regions for dropdown"""
    rbm_regions = Designation.objects.values_list('rbm_region', flat=True).distinct().order_by('rbm_region')
    return Response({
        'status': 'success',
        'rbm_regions': list(rbm_regions)
    }, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([AllowAny])
def validate_designation(request):
    """Validate if employee_id belongs to selected RBM region"""
    employee_id = request.data.get('employee_id')
    rbm_region = request.data.get('rbm_region')
    
    if not employee_id or not rbm_region:
        return Response({
            'status': 'error',
            'message': 'Both employee_id and rbm_region are required'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        designation = Designation.objects.get(login_code=employee_id, rbm_region=rbm_region)
        return Response({
            'status': 'success',
            'valid': True,
            'message': f'{employee_id} belongs to {rbm_region}'
        })
    except Designation.DoesNotExist:
        return Response({
            'status': 'error',
            'valid': False,
            'message': f'{employee_id} does not belong to {rbm_region}'
        })

# ------------------------------------------------------------------------------
# Doctors (base)
# ------------------------------------------------------------------------------

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
# #@ratelimit(key='ip', rate='20/m', method='POST', block=True)
def add_doctor(request):
    # Validate uploaded files
    if 'image' in request.FILES:
        try:
            validate_file_upload(request.FILES['image'])
        except ValidationError as e:
            return Response({'status': 'error', 'errors': {'image': str(e)}}, status=status.HTTP_400_BAD_REQUEST)

    serializer = DoctorSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({'status': 'success', 'data': serializer.data}, status=status.HTTP_201_CREATED)
    return Response({'status': 'error', 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

# ------------------------------------------------------------------------------
# Video generation (DoctorVideo)
# ------------------------------------------------------------------------------

class VideoGenViewSet(viewsets.ModelViewSet):
    queryset = DoctorVideo.objects.all()
    serializer_class = DoctorVideoSerializer

    # --- DYNAMIC time-duration parsing brought in from later versions ---
    def parse_time_duration(self, time_duration_str):
        """
        Parse a string like "2-6,65-70" into [(2,6),(65,70)].
        Validates order and non-negative starts.
        """
        if not time_duration_str or not time_duration_str.strip():
            raise ValueError("Time duration cannot be empty")
        try:
            slots = []
            for time_range in time_duration_str.strip().split(','):
                start_str, end_str = time_range.strip().split('-')
                start_time = int(start_str.strip())
                end_time = int(end_str.strip())
                if start_time >= end_time:
                    raise ValueError(f"Start time ({start_time}) must be less than end time ({end_time})")
                if start_time < 0:
                    raise ValueError(f"Start time ({start_time}) cannot be negative")
                slots.append((start_time, end_time))
            return slots
        except ValueError as e:
            if "not enough values" in str(e) or "too many values" in str(e):
                raise ValueError(
                    f"Invalid time duration format: '{time_duration_str}'. "
                    "Use '10-15' or '10-15,46-50'"
                )
            raise
        except Exception as e:
            raise ValueError(f"Error parsing time duration '{time_duration_str}': {str(e)}")

    def perform_create(self, serializer):
        """
        Save DoctorVideo only - video generation disabled
        """
        logger.info("VideoGenViewSet.perform_create called")
        doctor = serializer.save()
        logger.info(f"Doctor saved: {doctor.name} (ID: {doctor.id})")
        # Video generation disabled - only saving doctor data



class DoctorVideoViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = DoctorVideoSerializer
    pagination_class = Pagination_class
    permission_classes = [AllowAny]

    def get_queryset(self):
        employee_id = self.request.query_params.get('employee_id')
        user_type = self.request.query_params.get('user_type')
        search = self.request.query_params.get('search', '')
        specialization = self.request.query_params.get('specialization', '')

        if not employee_id:
            return DoctorVideo.objects.none()

        # Optimized for image-heavy queries
        if user_type in ["Admin", "SuperAdmin"]:
            doctors = DoctorVideo.objects.select_related('employee').prefetch_related('image_contents__template')
        else:
            try:
                employee = Employee.objects.get(employee_id=employee_id)
                doctors = DoctorVideo.objects.filter(employee=employee).select_related('employee').prefetch_related('image_contents__template')
            except Employee.DoesNotExist:
                return DoctorVideo.objects.none()

        # Apply filters
        if search:
            doctors = doctors.filter(
                Q(name__icontains=search) |
                Q(clinic__icontains=search) |
                Q(mobile_number__icontains=search)
            )
        if specialization:
            doctors = doctors.filter(specialization__iexact=specialization)

        return doctors.order_by('-created_at')



@api_view(['POST'])
@parser_classes([MultiPartParser])
# #@ratelimit(key='ip', rate='5/h', method='POST', block=True)
def bulk_upload_employees(request):
    excel_file = request.FILES.get('file')
    if not excel_file:
        return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        df = pd.read_excel(excel_file)
        required = {'first_name', 'last_name', 'email', 'phone', 'department', 'date_joined'}
        if not required.issubset(df.columns):
            return Response({'error': f'Missing required columns: {required - set(df.columns)}'}, status=status.HTTP_400_BAD_REQUEST)

        created, skipped, errors = 0, 0, []

        def generate_employee_id(first_name, last_name):
            base = f"EMP{first_name[0].upper()}{last_name[0].upper()}"
            while True:
                suffix = ''.join(random.choices(string.digits, k=4))
                emp_id = f"{base}{suffix}"
                if not Employee.objects.filter(employee_id=emp_id).exists():
                    return emp_id

        for index, row in df.iterrows():
            row_number = index + 2
            first_name = str(row.get('first_name', '')).strip()
            last_name = str(row.get('last_name', '')).strip()
            email = str(row.get('email', '')).strip()
            phone = str(row.get('phone', '')).strip()

            if not first_name or not last_name or not email:
                errors.append({'row': row_number, 'error': 'Missing first name, last name, email or phone'})
                skipped += 1
                continue

            employee_id = str(row.get('employee_id')).strip()
            data = {
                'employee_id': employee_id or generate_employee_id(first_name, last_name),
                'first_name': first_name,
                'last_name': last_name,
                'email': email,
                'phone': phone,
                'department': row.get('department', '') or 'Unknown',
                'date_joined': row.get('date_joined') if not pd.isna(row.get('date_joined')) else datetime.today().date()
            }
            serializer = EmployeeSerializer(data=data)
            if serializer.is_valid():
                try:
                    serializer.save()
                    created += 1
                except IntegrityError:
                    errors.append({'row': row_number, 'error': 'Duplicate email or other unique constraint'})
                    skipped += 1
            else:
                errors.append({'row': row_number, 'error': serializer.errors})
                skipped += 1

        return Response({'created': created, 'skipped': skipped, 'errors': errors}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# class DoctorListByEmployee(APIView):
#     permission_classes = [AllowAny]

#     def get(self, request):
#         employee_id = request.GET.get('employee_id')
#         user_type = request.GET.get('user_type')
#         search = request.GET.get('search', '')
#         specialization = request.GET.get('specialization', '')

#         if not employee_id:
#             return Response({"detail": "employee_id is required."}, status=400)

#         try:
#             employee = Employee.objects.get(employee_id=employee_id)
#         except Employee.DoesNotExist:
#             return Response({"detail": "Employee not found."}, status=404)

#         # Modified logic to include shared doctors
#         if user_type in ["Admin", "SuperAdmin"]:
#             doctors = DoctorVideo.objects.all()
#         else:
#             # Get doctors owned by this employee
#             owned_doctors = DoctorVideo.objects.filter(employee=employee)

#             # Get doctors used by this employee (from usage history)
#             used_doctor_ids = DoctorUsageHistory.objects.filter(
#                 employee=employee,
#                 content_type='image'
#             ).values_list('doctor_id', flat=True).distinct()

#             used_doctors = DoctorVideo.objects.filter(id__in=used_doctor_ids)

#             # Combine both querysets
#             doctors = (owned_doctors | used_doctors).distinct()

#         # Apply search filters
#         if search:
#             doctors = doctors.filter(
#                 Q(name__icontains=search) |
#                 Q(clinic__icontains=search)
#             )
#         if specialization:
#             doctors = doctors.filter(specialization__iexact=specialization)

#         doctors = doctors.order_by('-created_at')
#         paginator = Pagination_class()
#         page = paginator.paginate_queryset(doctors, request)
#         serializer = DoctorVideoSerializer(page, many=True)
#         return paginator.get_paginated_response(serializer.data)


class DoctorListByEmployee(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        employee_id = request.GET.get('employee_id')
        user_type = request.GET.get('user_type')
        search = request.GET.get('search', '')
        specialization = request.GET.get('specialization', '')

        if not employee_id:
            return Response({"detail": "employee_id is required."}, status=400)

        try:
            employee = Employee.objects.get(employee_id=employee_id)
            print(f"Employee found: {employee.first_name} {employee.last_name}")
        except Employee.DoesNotExist:
            return Response({"detail": "Employee not found."}, status=404)

        # MODIFIED - Employee-specific doctors only (shared doctor functionality commented out)
        if user_type in ["Admin", "SuperAdmin"]:
            doctors = DoctorVideo.objects.all()
            print(f"Admin mode: returning all {doctors.count()} doctors")
        else:
            # Only get doctors owned by this employee (no shared doctors)
            doctors = DoctorVideo.objects.filter(employee=employee)
            print(f"Doctors owned by {employee_id}: {doctors.count()}")

            # COMMENTED OUT - Shared doctor functionality
            # used_doctor_ids = DoctorUsageHistory.objects.filter(
            #     employee=employee,
            #     content_type='image'
            # ).values_list('doctor_id', flat=True).distinct()
            # used_doctors = DoctorVideo.objects.filter(id__in=used_doctor_ids)
            # doctors = (owned_doctors | used_doctors).distinct()

        # Apply search filters
        if search:
            doctors = doctors.filter(
                Q(name__icontains=search) |
                Q(clinic__icontains=search)
            )
        if specialization:
            doctors = doctors.filter(specialization__iexact=specialization)

        doctors = doctors.order_by('-created_at')
        paginator = Pagination_class()
        page = paginator.paginate_queryset(doctors, request)
        serializer = DoctorVideoSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)


class DoctorVideoListView(APIView):
    def get(self, request):
        employee_id = request.GET.get('employee_id')
        if not employee_id:
            return Response({"detail": "employee_id is required."}, status=400)
        try:
            employee = Employee.objects.get(employee_id=employee_id)
        except Employee.DoesNotExist:
            return Response({"detail": "Employee not found for this employee_id."}, status=404)
        doctors = DoctorVideo.objects.filter(employee=employee).order_by('-created_at')
        paginator = Pagination_class()
        page = paginator.paginate_queryset(doctors, request)
        serializer = DoctorVideoSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)

# ------------------------------------------------------------------------------
# Manual single doctor generation (kept as-is; uses sync generation)
# ------------------------------------------------------------------------------

class DoctorVideoGeneration(APIView):
    def post(self, request):
        pass


class CustomTokenRefreshView(TokenRefreshView):
    def post(self, request, *args, **kwargs):
        refresh_token = request.data.get('refresh', None)
        if not refresh_token:
            return Response({'status': 'error', 'message': 'Refresh token is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            refresh = RefreshToken(refresh_token)
            new_refresh_token = str(refresh)
            new_access_token = str(refresh.access_token)
            return Response({
                'status': 'success',
                'message': 'Token refreshed successfully',
                'access': new_access_token,
                'refresh': new_refresh_token,
                'access_token_exp': refresh.access_token.payload['exp'],
            }, status=status.HTTP_200_OK)
        except Exception:
            return Response({'status': 'error', 'message': 'Invalid refresh token'}, status=status.HTTP_401_UNAUTHORIZED)

def generate_video_for_doctor(doctor):
    """Simple hardcoded-template generation path (kept for backward compatibility)."""
    logger.info(f"Generating video for doctor {doctor.name}")
    try:
        if not doctor.image:
            logger.warning(f"No image for doctor {doctor.name}, skipping video.")
            return

        main_video_path = os.path.join(settings.MEDIA_ROOT, "Health111.mp4")
        image_path = doctor.image.path
        output_dir = os.path.join(settings.MEDIA_ROOT, "output")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{doctor.id}_output.mp4")

        # Reuse the sync path from VideoGenViewSet for consistent visuals
        VideoGenViewSet().generate_custom_video(
            main_video_path=main_video_path,
            image_path=image_path,
            name=doctor.name,
            clinic=doctor.clinic,
            city=doctor.city,
            specialization_key=getattr(doctor, "specialization_key", doctor.specialization),
            state=doctor.state,
            output_path=output_path
        )

        with open(output_path, 'rb') as f:
            doctor.output_video.save(f"{doctor.id}_output.mp4", File(f), save=True)

        BASE_URL = "https://api.videomaker.digielvestech.in/"
        doctor.output_video_url = f"{BASE_URL}{doctor.output_video.url}"
        doctor.save()
        logger.info(f"Video generated and saved for doctor {doctor.name}")

    except Exception as e:
        logger.error(f"Error generating video for doctor {doctor.name}: {e}")

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
# #@ratelimit(key='ip', rate='5/h', method='POST', block=True)
def bulk_upload_doctors(request):
    excel_file = request.FILES.get('file')
    if not excel_file:
        return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        df = pd.read_excel(excel_file)
        required_columns = {'name', 'clinic', 'city', 'specialization', 'state'}
        if not required_columns.issubset(df.columns):
            return Response({'error': f'Missing required columns: {required_columns - set(df.columns)}'}, status=status.HTTP_400_BAD_REQUEST)

        created, skipped, errors = 0, 0, []
        for index, row in df.iterrows():
            row_number = index + 2
            name = str(row.get('name', '')).strip()
            clinic = str(row.get('clinic', '')).strip()
            city = str(row.get('city', '')).strip()
            specialization = str(row.get('specialization', '')).strip()
            state = str(row.get('state', '')).strip()

            if not name or not clinic or not city or not specialization or not state:
                skipped += 1
                errors.append({'row': row_number, 'error': 'Required fields are missing'})
                continue

            designation = str(row.get('designation', '')).strip()
            mobile_number = str(row.get('mobile_number', '')).strip()
            whatsapp_number = str(row.get('whatsapp_number', '')).strip()
            description = str(row.get('description', '')).strip()
            image_path = str(row.get('image_url', '')).strip()
            employee = row.get('emp_id')

            image_file = None
            if image_path and os.path.exists(image_path):
                image_file = File(open(image_path, 'rb'), name=os.path.basename(image_path))

            doctor_data = {
                'name': name,
                'clinic': clinic,
                'city': city,
                'specialization': specialization,
                'state': state,
                'image': image_file,
                'designation': designation,
                'mobile_number': mobile_number,
                'whatsapp_number': whatsapp_number,
                'description': description,
                'employee': employee,
            }

            serializer = DoctorSerializer(data=doctor_data)
            if serializer.is_valid():
                try:
                    doctor = serializer.save()
                    generate_video_for_doctor(doctor)
                    created += 1
                except IntegrityError:
                    skipped += 1
                    errors.append({'row': row_number, 'error': 'Integrity error during save'})
            else:
                skipped += 1
                errors.append({'row': row_number, 'error': serializer.errors})

        return Response({'created': created, 'skipped': skipped, 'errors': errors}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ------------------------------------------------------------------------------
# Export/Counts
# ------------------------------------------------------------------------------

class DoctorVideoExportExcelView(APIView):
    def get(self, request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Doctor Videos"
        headers = [
            "Name", "Designation", "Clinic", "City", "State", "Image URL",
            "Specialization", "Mobile Number", "WhatsApp Number", "Description", "Template Name",
            "Output Video URL", "Created At", "Employee ID", "Employee Name", "RBM Name"
        ]
        sheet.append(headers)
        doctor_videos = DoctorVideo.objects.all()
        for video in doctor_videos:
            # DoctorOutputVideo model removed - use basic output_video field only
            if video.output_video:
                output_video_url = request.build_absolute_uri(video.output_video.url)
                template_name = ""
            else:
                output_video_url = ""
                template_name = ""
            rbm_name = (
                f"{video.employee.rbm.first_name} {video.employee.rbm.last_name}"
                if video.employee and getattr(video.employee, 'rbm', None) else ""
            )
            sheet.append([
                video.name, video.designation, video.clinic, video.city, video.state,
                request.build_absolute_uri(video.image.url) if video.image else "",
                video.specialization, video.mobile_number, video.whatsapp_number, video.description,
                template_name, output_video_url,
                video.created_at.strftime('%Y-%m-%d %H:%M:%S') if video.created_at else "",
                video.employee.employee_id if video.employee else "",
                f"{video.employee.first_name} {video.employee.last_name}" if video.employee else "",
                rbm_name,
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=doctor_videos.xlsx'
        workbook.save(response)
        return response

class EmployeeExportExcelView(APIView):
    def get(self, request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Employees"
        headers = ["Employee ID", "First Name", "Last Name", "Email", "Phone", "Department", "Date Joined", "User Type", "Status"]
        sheet.append(headers)
        for emp in Employee.objects.all():
            sheet.append([
                emp.employee_id, emp.first_name, emp.last_name or "", emp.email or "", emp.phone or "",
                emp.department or "", emp.date_joined.strftime('%Y-%m-%d') if emp.date_joined else "",
                emp.user_type, "Active" if emp.status else "Inactive"
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=employees.xlsx'
        workbook.save(response)
        return response

@api_view(['GET'])
def total_employee_count(request):
    return Response({"total_employees": Employee.objects.count()}, status=status.HTTP_200_OK)

@api_view(['GET'])
def todays_active_employees(request):
    today = timezone.now().date()
    employees = Employee.objects.filter(login_history__login_time__date=today).distinct()
    count = employees.count()
    data = [{
        'employee_id': emp.employee_id,
        'name': f"{emp.first_name} {emp.last_name}",
        'email': emp.email,
        'department': emp.department,
        'user_type': emp.user_type,
    } for emp in employees]
    return Response({'date': str(today), 'active_employee_count': count, 'active_employees': data})

class TodaysActiveEmployeeExcelExport(APIView):
    def get(self, request):
        today = timezone.now().date()
        employees = Employee.objects.filter(login_history__login_time__date=today).distinct()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Today's Active Employees"
        headers = ["Employee ID", "First Name", "Last Name", "Email", "Phone", "Department", "Date Joined", "User Type", "Status"]
        sheet.append(headers)
        for emp in employees:
            sheet.append([
                emp.employee_id, emp.first_name, emp.last_name or "", emp.email or "", emp.phone or "",
                emp.department or "", emp.date_joined.strftime('%Y-%m-%d') if emp.date_joined else "",
                emp.user_type, "Active" if emp.status else "Inactive"
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=active_employees_{today}.xlsx'
        workbook.save(response)
        return response

@api_view(['GET'])
def doctors_with_output_video_count(request):
    count = DoctorVideo.objects.filter(Q(output_video__isnull=False) & ~Q(output_video='')).count()
    return Response({"doctor_with_output_video_count": count}, status=status.HTTP_200_OK)

@api_view(['GET'])
def doctors_with_output_video_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Doctor Videos"
    headers = [
        "Name", "Designation", "Clinic", "City", "State", "Image URL",
        "Specialization", "Mobile Number", "WhatsApp Number", "Description",
        "Output Video URL", "Created At", "Employee ID", "Employee Name", "RBM Name"
    ]
    sheet.append(headers)
    doctor_videos = DoctorVideo.objects.all()
    for video in doctor_videos:
        rbm_name = (
            f"{video.employee.rbm.first_name} {video.employee.rbm.last_name}"
            if video.employee and getattr(video.employee, 'rbm', None) else ""
        )
        sheet.append([
            video.name, video.designation, video.clinic, video.city, video.state,
            request.build_absolute_uri(video.image.url) if video.image else "",
            video.specialization, video.mobile_number, video.whatsapp_number, video.description,
            request.build_absolute_uri(video.output_video.url) if video.output_video else "",
            video.created_at.strftime('%Y-%m-%d %H:%M:%S') if video.created_at else "",
            video.employee.employee_id if video.employee else "",
            f"{video.employee.first_name} {video.employee.last_name}" if video.employee else "",
            rbm_name,
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=doctor_videos.xlsx'
    workbook.save(response)
    return response

@api_view(['GET'])
def doctors_count(request):
    return Response({"doctor_count": DoctorVideo.objects.count()}, status=status.HTTP_200_OK)

# ------------------------------------------------------------------------------
# Templates (Video)
# ------------------------------------------------------------------------------

class VideoTemplateAPIView(APIView):
    permission_classes = [AllowAny]  # Add this line
    def get(self, request, pk=None):
        if pk:
            template = get_object_or_404(VideoTemplates, pk=pk)
            serializer = VideoTemplatesSerializer(template)
        else:
            status_param = request.query_params.get('status')
            template_type = request.query_params.get('template_type', 'image') # <- NEW
            user_type = request.query_params.get('user_type')
            employee_id = request.query_params.get('employee_id')

            templates = VideoTemplates.objects.filter(template_type='image')# <- NEW

            # Apply user-based filtering
            if user_type and user_type not in ["Admin", "SuperAdmin"]:
                # Non-admin users see only public templates or their own
                if employee_id:
                    try:
                        employee = Employee.objects.get(employee_id=employee_id)
                        templates = templates.filter(
                            Q(is_public=True) | Q(created_by=employee)
                        )
                    except Employee.DoesNotExist:
                        templates = templates.filter(is_public=True)
                else:
                    templates = templates.filter(is_public=True)

            if status_param is not None:
                try:
                    status_bool = bool(strtobool(status_param))
                    templates = templates.filter(status=status_bool)
                except ValueError:
                    return Response({"error": "Invalid status value. Use true or false."},
                                    status=status.HTTP_400_BAD_REQUEST)

            templates = templates.order_by('-created_at')
            serializer = VideoTemplatesSerializer(templates, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
            serializer = VideoTemplatesSerializer(data=request.data)
            if serializer.is_valid():
                # Set the creator
                employee_id = request.data.get('employee_id') or getattr(request.user, 'username', None)
                if employee_id:
                    try:
                        employee = Employee.objects.get(employee_id=employee_id)
                        serializer.save(created_by=employee)
                    except Employee.DoesNotExist:
                        serializer.save()
                else:
                    serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        template = get_object_or_404(VideoTemplates, pk=pk)
        serializer = VideoTemplatesSerializer(template, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        template = get_object_or_404(VideoTemplates, pk=pk)
        template.delete()
        return Response({"detail": "Deleted successfully."}, status=status.HTTP_204_NO_CONTENT)

# ------------------------------------------------------------------------------
# On-demand generation (Video) with dynamic time slots
# ------------------------------------------------------------------------------

class GenerateDoctorOutputVideoView(APIView):
    pass


# Employee updates from Excel
# ------------------------------------------------------------------------------

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def update_employees_from_excel(request):
    excel_file = request.FILES.get('file')
    if not excel_file:
        return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        df = pd.read_excel(excel_file)
        updated, not_found = 0, 0
        for _, row in df.iterrows():
            employee_id = str(row['id']).strip()
            department = str(row['department']).strip()
            city = str(row['city']).strip().title()
            try:
                employee = Employee.objects.get(employee_id=employee_id)
                employee.department = department
                employee.city = city
                employee.save()
                updated += 1
            except Employee.DoesNotExist:
                not_found += 1
        return Response({'status': 'success', 'updated': updated, 'not_found': not_found})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class TemplateWiseVideoCountView(APIView):
    def get(self, request):
        template_type = request.GET.get('template_type', 'video')  # Add this parameter

        if template_type == 'image':
            # Count ImageContent instead of DoctorOutputVideo
            template_counts = ImageContent.objects.values("template__id", "template__name").annotate(content_count=Count("id")).order_by("-content_count")
        else:
            # Video functionality disabled - return empty results
            template_counts = []

        data = [{
            "template_id": item["template__id"],
            "template_name": item["template__name"],
            "video_count": item.get("video_count") or item.get("content_count", 0)
        } for item in template_counts if item["template__id"] is not None]

        return Response(data, status=status.HTTP_200_OK)

# ------------------------------------------------------------------------------
# Image Templates & Image Generation
# ------------------------------------------------------------------------------

class ImageTemplateAPIView(APIView):
    """Handle image templates separately from video templates"""
    permission_classes = [AllowAny]
    def get(self, request, pk=None):
        if pk:
            template = get_object_or_404(VideoTemplates, pk=pk, template_type='image')
            serializer = ImageTemplateSerializer(template, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            status_param = request.query_params.get('status')
            templates = VideoTemplates.objects.filter(template_type='image')
            if status_param is not None:
                try:
                    status_bool = bool(strtobool(status_param))
                    templates = templates.filter(status=status_bool)
                except ValueError:
                    return Response({"error": "Invalid status value. Use true or false."}, status=status.HTTP_400_BAD_REQUEST)
            templates = templates.order_by('-created_at')
            serializer = ImageTemplateSerializer(templates, many=True, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = ImageTemplateSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        template = get_object_or_404(VideoTemplates, pk=pk, template_type='image')
        serializer = ImageTemplateSerializer(template, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        template = get_object_or_404(VideoTemplates, pk=pk, template_type='image')
        template.delete()
        return Response({"detail": "Template deleted successfully."}, status=status.HTTP_204_NO_CONTENT)

# @method_decorator(ratelimit(key='ip', rate='5/m', method='POST', block=True), name='post')

class GenerateImageContentView(APIView):
    """Generate image with text overlay - DoctorVideo only"""
    # #throttle_classes = []
    permission_classes = [AllowAny]
    @monitor_resources
    def post(self, request):
        # Check system resources before processing
        memory_percent = psutil.virtual_memory().percent
        if memory_percent > 95:
            return Response({
                "error": "System under high load, please try again later"
            }, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        try:
            return self._process_request(request)
        except ValidationError as e:
            logger.error(f"Validation error in image generation: {e}")
            return Response({
                "error": "Invalid input data",
                "details": str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
        except VideoTemplates.DoesNotExist:
            logger.error(f"Template not found: {request.data.get('template_id')}")
            return Response({
                "error": "Template not found"
            }, status=status.HTTP_404_NOT_FOUND)
        except DoctorVideo.DoesNotExist:
            logger.error(f"Doctor not found: {request.data.get('doctor_id')}")
            return Response({
                "error": "Doctor not found"
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Unexpected error in image generation: {e}", exc_info=True)
            return Response({
                "error": "Image generation failed",
                "message": "Please try again later"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _process_request(self, request):
        # Add security check - get from request data, not localStorage
        employee_id = request.data.get("employee_id")
        user_type = request.data.get("user_type", "Employee")

        # Validate uploaded files first
        if 'doctor_image' in request.FILES:
            try:
                validate_file_upload(request.FILES['doctor_image'])
            except ValidationError as e:
                return Response({"error": f"File validation failed: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        template_id = request.data.get("template_id")
        doctor_id = request.data.get("doctor_id")
        mobile = request.data.get("mobile")
        name = request.data.get("name")
        content_data = request.data.get("content_data", {})
        selected_brand_ids = request.data.get("selected_brands", [])

        # Security: Ensure employee can only create content for themselves (unless admin)
        if doctor_id and user_type not in ["Admin", "SuperAdmin"]:
            try:
                doctor = DoctorVideo.objects.get(id=doctor_id)
                if doctor.employee.employee_id != employee_id:
                    return Response({"error": "You can only generate content for your own doctors"}, status=status.HTTP_403_FORBIDDEN)
            except DoctorVideo.DoesNotExist:
                pass  # Will be handled later

        if not isinstance(selected_brand_ids, list) or not all(isinstance(bid, int) for bid in selected_brand_ids):
            return Response({"error": "selected_brands must be a list of brand IDs."}, status=status.HTTP_400_BAD_REQUEST)

        if len(selected_brand_ids) > 10:
            return Response({"error": "You can select up to 10 brands only."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            template = VideoTemplates.objects.select_related().get(id=template_id, template_type='image')
        except VideoTemplates.DoesNotExist:
            return Response({"error": "Image template not found."}, status=status.HTTP_404_NOT_FOUND)

        print(f"🔍 Template brand_area_settings: {template.brand_area_settings}")
        print(f"🔍 Selected brand IDs: {selected_brand_ids}")


        # Scenario 1: Existing DoctorVideo by ID
        is_new_doctor = False
        if doctor_id:
            try:
                doctor_video = DoctorVideo.objects.select_related('employee').get(id=doctor_id)
            except DoctorVideo.DoesNotExist:
                return Response({"error": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)
        else:
            # Scenario 2: Employee-specific doctor handling
            try:
                employee_id = request.data.get("employee_id")
                if employee_id:
                    employee = Employee.objects.get(employee_id=employee_id)
                else:
                    return Response({"error": "employee_id is required."}, status=status.HTTP_400_BAD_REQUEST)
            except Employee.DoesNotExist:
                return Response({"error": f"Employee {employee_id} not found."}, status=status.HTTP_404_NOT_FOUND)
            
            # Check if THIS EMPLOYEE already has this doctor
            doctor_video = DoctorVideo.objects.select_related('employee').filter(
                mobile_number=mobile, 
                employee=employee
            ).first()
            
            if doctor_video:
                # Employee found their own existing doctor - UPDATE instead of create
                clinic = content_data.get("doctor_clinic") or request.data.get("clinic", doctor_video.clinic)
                city = content_data.get("doctor_city") or request.data.get("city", doctor_video.city)
                specialization = content_data.get("doctor_specialization") or request.data.get("specialization", doctor_video.specialization)
                state = content_data.get("doctor_state") or request.data.get("state", doctor_video.state)
                
                # Update existing doctor with new data (allow name changes)
                doctor_video.name = name  # Name can be updated
                doctor_video.clinic = clinic
                doctor_video.city = city  
                doctor_video.specialization = specialization
                doctor_video.specialization_key = specialization
                doctor_video.state = state
                
                # Handle image upload for existing doctor
                uploaded_image = request.FILES.get('doctor_image')
                if uploaded_image:
                    doctor_video.image = uploaded_image
                
                doctor_video.save()
                is_new_doctor = False
                
            else:
                # Check if OTHER employees have this doctor (for auto-population)
                existing_doctor = DoctorVideo.objects.filter(mobile_number=mobile).first()
                
                if existing_doctor:
                    # Auto-populate from existing doctor but create new record
                    clinic = content_data.get("doctor_clinic") or request.data.get("clinic", existing_doctor.clinic)
                    city = content_data.get("doctor_city") or request.data.get("city", existing_doctor.city) 
                    specialization = content_data.get("doctor_specialization") or request.data.get("specialization", existing_doctor.specialization)
                    state = content_data.get("doctor_state") or request.data.get("state", existing_doctor.state)
                else:
                    # Completely new doctor
                    clinic = content_data.get("doctor_clinic") or request.data.get("clinic", "Unknown Clinic")
                    city = content_data.get("doctor_city") or request.data.get("city", "Unknown City")
                    specialization = content_data.get("doctor_specialization") or request.data.get("specialization", "General Medicine")
                    state = content_data.get("doctor_state") or request.data.get("state", "Unknown State")

                # Create new doctor for this employee
                doctor_video = DoctorVideo.objects.create(
                    name=name,
                    clinic=clinic,
                    city=city,
                    specialization=specialization,
                    specialization_key=specialization,
                    state=state,
                    mobile_number=mobile,
                    whatsapp_number=mobile,
                    description=request.data.get("description", ""),
                    designation=request.data.get("designation", ""),
                    employee=employee
                )
                uploaded_image = request.FILES.get('doctor_image')
                if uploaded_image:
                    doctor_video.image = uploaded_image
                    doctor_video.save()
                is_new_doctor = True

        if not template.template_image or not template.template_image.path:
            return Response({"error": "Template does not have an image file."}, status=status.HTTP_400_BAD_REQUEST)

        # Verify template image file exists
        if not os.path.exists(template.template_image.path):
            return Response({"error": "Template image file not found."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Use background processing for images

            # print(f"About to call generate_image_with_text with brand IDs: {selected_brand_ids}")
            # output_path = self.generate_image_with_text(template, content_data, doctor_video, selected_brand_ids)
            # image_content = ImageContent.objects.create(
            #     template=template,
            #     doctor=doctor_video,
            #     content_data=content_data
            # )
            # with open(output_path, 'rb') as f:
            #     image_content.output_image.save(f"generated_{image_content.id}.png", File(f), save=True)
            # os.remove(output_path)

            # serializer = ImageContentSerializer(image_content, context={'request': request})
            # resp = serializer.data

            # Use background processing for images
            from .tasks import generate_image_async

            task = generate_image_async.delay(
                template_id=template_id,
                doctor_id=doctor_video.id,
                content_data=content_data,
                selected_brand_ids=selected_brand_ids,
                current_employee_id=employee_id  # Add this line
            )

            # Return task ID to frontend
            resp = {
                "status": "processing",
                "task_id": str(task.id),
                "message": "Image generation started"
            }



            resp['doctor_info'] = {
                'id': doctor_video.id,
                'name': doctor_video.name,
                'clinic': doctor_video.clinic,
                'mobile': doctor_video.mobile_number,
                'has_image': bool(doctor_video.image),
                'is_new_doctor': is_new_doctor
            }
            if selected_brand_ids:
                # Use single query instead of individual lookups
                selected_brands = list(Brand.objects.filter(id__in=selected_brand_ids).values('id', 'name', 'category'))
                resp['selected_brands'] = selected_brands
            else:
                resp['selected_brands'] = []


            return Response(resp, status=status.HTTP_201_CREATED)
        except Exception as e:
            # Handle database connection issues
            if "too many clients" in str(e) or "connection" in str(e).lower():
                logger.error(f"Database connection issue: {e}")
                return Response({
                    "error": "System temporarily unavailable",
                    "retry_after": 30
                }, status=status.HTTP_503_SERVICE_UNAVAILABLE)

            logger.error(f"Image generation failed: {e}")
            logger.error(f"Template ID: {template.id}, Template has image: {bool(template.template_image)}")
            if template.template_image:
                logger.error(f"Template image path: {template.template_image.path}")
            return Response({"error": "Image generation failed.", "details": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    #

    def generate_image_with_text(self, template, content_data, doctor, selected_brand_ids=None):
        logger.info(f"Starting image generation for template {template.id}")

        if not template.template_image or not template.template_image.path:
            raise Exception("Template image path is None or empty")

        if not os.path.exists(template.template_image.path):
            raise Exception(f"Template image file does not exist: {template.template_image.path}")

        # Memory-efficient image processing with explicit cleanup
        template_image = None
        draw = None
        brand_images = []  # Track all opened images for cleanup

        # Make brand_images accessible to render_brands_in_area method
        self._current_brand_images = brand_images
        try:
            template_image = Image.open(template.template_image.path)


            # Convert to RGB if needed to reduce memory
            if template_image.mode not in ('RGB', 'RGBA'):
                template_image = template_image.convert('RGB')


            draw = ImageDraw.Draw(template_image)

            logger.info("Getting text positions...")
            positions = template.text_positions or {}
            logger.info(f"Text positions: {positions}")

            positions = template.text_positions or {}
            # Font mapping for PIL
            FONT_MAP = {
                'Arial': 'arial.ttf',
                'Times New Roman': 'times.ttf',
                'Helvetica': 'arial.ttf',
                'Georgia': 'georgia.ttf',
                'Verdana': 'verdana.ttf',
                'Impact': 'impact.ttf',
                'Comic Sans MS': 'comic.ttf',
                'Dancing Script': 'DancingScript-Regular.ttf',
                'Great Vibes': 'GreatVibes-Regular.ttf',
                'Pacifico': 'Pacifico-Regular.ttf',
                'Allura': 'Allura-Regular.ttf',
                'Alex Brush': 'AlexBrush-Regular.ttf'
            }

            def get_font(font_family, font_size, font_weight='normal', font_style='normal'):
                base_font = FONT_MAP.get(font_family, 'arial.ttf')

                # For cursive fonts, use full path
                if font_family in ['Dancing Script', 'Great Vibes', 'Pacifico', 'Allura', 'Alex Brush']:
                    font_path = os.path.join(settings.BASE_DIR, "fonts", base_font)
                else:
                    font_path = base_font

                # Handle font weight (bold) - only for non-cursive fonts
                if font_weight == 'bold' and font_family not in ['Dancing Script', 'Great Vibes', 'Pacifico', 'Allura', 'Alex Brush']:
                    bold_font = font_path.replace('.ttf', 'bd.ttf')
                    try:
                        font = ImageFont.truetype(bold_font, font_size)
                    except:
                        try:
                            font = ImageFont.truetype(font_path, font_size)
                        except:
                            font = ImageFont.load_default()
                else:
                    try:
                        font = ImageFont.truetype(font_path, font_size)
                    except:
                        try:
                            font = ImageFont.truetype("arial.ttf", font_size)
                        except:
                            font = ImageFont.load_default()

                return font
            ##
            # Combine city and state with comma if both exist
            city_state = []
            if content_data.get('doctor_city', doctor.city):
                city_state.append(content_data.get('doctor_city', doctor.city))
            if content_data.get('doctor_state', doctor.state):
                city_state.append(content_data.get('doctor_state', doctor.state))
            city_state_combined = ', '.join(city_state)

            all_text_data = {
                'name': content_data.get('doctor_name', doctor.name),
                'city': city_state_combined,  # Combined city, state
                'specialization': content_data.get('doctor_specialization', doctor.specialization),
                'mobile': doctor.mobile_number,
                'customText': template.custom_text or '',
            }

            # DOCTOR FIELDS RESPONSIVE CENTER ALIGNMENT (removed 'state' since it's now combined with city)
            doctor_fields = ['name', 'specialization', 'city']
            doctor_text_data = {field: all_text_data[field] for field in doctor_fields if field in all_text_data and all_text_data[field] and field in positions}

            if doctor_text_data:
                print(f"🎯 Processing doctor fields for center alignment: {list(doctor_text_data.keys())}")

                # Calculate text widths for center alignment
                field_widths = {}
                field_fonts = {}

                for field_name, text_value in doctor_text_data.items():
                    pos = positions[field_name]
                    font_size = int(pos.get('fontSize', 40)) * 7
                    font_weight = pos.get('fontWeight', 'normal')
                    font_family = pos.get('fontFamily', 'Arial')
                    font_style = pos.get('fontStyle', 'normal')

                    styled_font = get_font(font_family, font_size, font_weight, font_style)
                    field_fonts[field_name] = styled_font

                    # Calculate text width using textbbox
                    bbox = draw.textbbox((0, 0), str(text_value), font=styled_font)
                    text_width = bbox[2] - bbox[0]
                    field_widths[field_name] = text_width
                    print(f"🎯 Field {field_name}: '{text_value}' = {text_width}px wide")

                # Find the widest text to base centering on
                max_width = max(field_widths.values())
                widest_field = max(field_widths, key=field_widths.get)
                print(f"🎯 Widest field: {widest_field} ({max_width}px)")

                # Get base position from widest field
                # Calculate the visual center point of the template
                template_center_x = template_image.width // 2

                # Calculate center-aligned positions for all doctor fields
                for field_name, text_value in doctor_text_data.items():
                    pos = positions[field_name]
                    field_width = field_widths[field_name]
                    styled_font = field_fonts[field_name]

                    # Center each field based on template center, not relative to other fields
                    centered_x = template_center_x - (field_width // 2)
                    y_pos = int(pos['y'])  # Keep original Y position
                    print(f"🎯 Rendering {field_name} at centered position ({centered_x}, {y_pos})")

                    # Apply styling
                    color = pos.get('color', 'black')
                    font_style = pos.get('fontStyle', 'normal')

                    # Text shadow
                    text_shadow = pos.get('textShadow', 'none')
                    if text_shadow != 'none':
                        shadow_info = parse_css_shadow(text_shadow)
                        if shadow_info:
                            draw.text(
                                (centered_x + shadow_info['offset_x'], y_pos + shadow_info['offset_y']),
                                str(text_value), fill=shadow_info['color'], font=styled_font
                            )

                    # Render text with center alignment
                    # For cursive fonts, no need for italic simulation - they're naturally cursive
                    # For non-cursive fonts with italic style, apply simulation
                    if font_style in ['italic', 'oblique'] and pos.get('fontFamily', 'Arial') not in ['Dancing Script', 'Great Vibes', 'Pacifico', 'Allura', 'Alex Brush']:
                        for offset in range(3):
                            draw.text((centered_x + offset, y_pos), str(text_value), fill=color, font=styled_font)
                    else:
                        draw.text((centered_x, y_pos), str(text_value), fill=color, font=styled_font)

            # Handle custom text separately (not part of doctor info centering)
            # Handle custom text separately (not part of doctor info centering) - FIXED POSITION
            if 'customText' in all_text_data and 'customText' in positions and all_text_data['customText']:
                field_name = 'customText'
                text_value = all_text_data[field_name]
                pos = positions[field_name]
                font_size = int(pos.get('fontSize', 40))
                color = pos.get('color', 'black')
                font_weight = pos.get('fontWeight', 'normal')
                font_family = pos.get('fontFamily', 'Arial')
                font_style = pos.get('fontStyle', 'normal')

                styled_font = get_font(font_family, font_size, font_weight, font_style)

                # Use original fixed position for custom text (no responsive centering)
                x_pos = int(pos['x'])
                y_pos = int(pos['y'])

                text_shadow = pos.get('textShadow', 'none')
                if text_shadow != 'none':
                    shadow_info = parse_css_shadow(text_shadow)
                    if shadow_info:
                        draw.text(
                            (x_pos + shadow_info['offset_x'], y_pos + shadow_info['offset_y']),
                            str(text_value), fill=shadow_info['color'], font=styled_font
                        )

                if font_style in ['italic', 'oblique'] and font_family not in ['Dancing Script', 'Great Vibes', 'Pacifico', 'Allura', 'Alex Brush']:
                    for offset in range(3):
                        draw.text((x_pos + offset, y_pos), str(text_value), fill=color, font=styled_font)
                else:
                    draw.text((x_pos, y_pos), str(text_value), fill=color, font=styled_font)
            # Optional doctor image overlay
            image_settings = None
            if content_data and 'imageSettings' in content_data:
                image_settings = content_data['imageSettings']
            elif template.text_positions and 'imageSettings' in template.text_positions:
                image_settings = template.text_positions['imageSettings']

            if image_settings and image_settings.get('enabled', False):
                try:
                    logger.info(f"Doctor image enabled. Doctor: {doctor.name}")
                    logger.info(f"Doctor image field: {doctor.image}")
                    logger.info(f"Doctor image path: {doctor.image.path if doctor.image else 'None'}")

                    img_x = int(image_settings.get('x', 400))
                    img_y = int(image_settings.get('y', 50))
                    img_width = int(image_settings.get('width', 150))
                    img_height = int(image_settings.get('height', 150))
                    img_fit = image_settings.get('fit', 'cover')
                    border_radius = int(image_settings.get('borderRadius', 0))
                    opacity = int(image_settings.get('opacity', 100))

                    if doctor.image and doctor.image.path and os.path.exists(doctor.image.path):
                        doctor_img = Image.open(doctor.image.path)
                        if img_fit == 'cover':
                            doctor_img = doctor_img.resize((img_width, img_height), Image.Resampling.LANCZOS)
                        elif img_fit == 'contain':
                            doctor_img.thumbnail((img_width, img_height), Image.Resampling.LANCZOS)
                        elif img_fit == 'stretch':
                            doctor_img = doctor_img.resize((img_width, img_height), Image.Resampling.LANCZOS)
                    else:
                        doctor_img = Image.new('RGB', (img_width, img_height), color='#4A90E2')
                        draw_placeholder = ImageDraw.Draw(doctor_img)
                        fsz = max(20, min(img_width, img_height) // 4)
                        try:
                            placeholder_font = ImageFont.truetype("arial.ttf", fsz)
                        except:  # noqa: E722
                            placeholder_font = ImageFont.load_default()
                        draw_placeholder.text((img_width//2, img_height//2), "DR", fill='white', font=placeholder_font, anchor="mm")

                    if border_radius > 0:
                        mask = Image.new('L', (doctor_img.width, doctor_img.height), 0)
                        mask_draw = ImageDraw.Draw(mask)
                        actual_radius = min(doctor_img.width, doctor_img.height) * border_radius // 200
                        mask_draw.rounded_rectangle([(0, 0), (doctor_img.width, doctor_img.height)], radius=actual_radius, fill=255)
                        if doctor_img.mode != 'RGBA':
                            doctor_img = doctor_img.convert('RGBA')
                        doctor_img.putalpha(mask)

                    if opacity < 100:
                        if doctor_img.mode != 'RGBA':
                            doctor_img = doctor_img.convert('RGBA')
                        alpha = doctor_img.split()[-1]
                        alpha = alpha.point(lambda p: int(p * opacity / 100))
                        doctor_img.putalpha(alpha)

                    if template_image.mode != 'RGBA':
                        template_image = template_image.convert('RGBA')
                    if doctor_img.mode == 'RGBA':
                        template_image.paste(doctor_img, (img_x, img_y), doctor_img)
                    else:
                        template_image.paste(doctor_img, (img_x, img_y))
                except Exception as e:
                    logger.error(f"Error compositing doctor image: {e}")

            #output_dir = os.path.join(settings.MEDIA_ROOT, "temp")

            output_dir = os.path.join(settings.MEDIA_ROOT, "temp")

            # Render brands in predefined area with smart layout
    # Render brands in predefined area with smart layout
            if selected_brand_ids is None:
                selected_brand_ids = content_data.get('selected_brands', [])

            print(f"🔍 Final selected_brand_ids after processing: {selected_brand_ids}")
            print(f"🔍 Template brand_area_settings: {template.brand_area_settings}")
            print(f"🔍 Brand area enabled: {template.brand_area_settings.get('enabled', False) if template.brand_area_settings else False}")
            logger.info(f"Template ID: {template.id}, Template name: {template.name}")
            logger.info(f"Brand area settings: {template.brand_area_settings}")
            logger.info(f"Selected brand IDs: {selected_brand_ids}")

            print(f"🔍 Checking condition: selected_brand_ids={bool(selected_brand_ids)}, has_area_settings={bool(template.brand_area_settings)}, area_enabled={template.brand_area_settings.get('enabled', False) if template.brand_area_settings else False}")

            if selected_brand_ids and template.brand_area_settings and template.brand_area_settings.get('enabled', False):
                print("🔍 CONDITION MET - About to render brands")
                print(f"🔍 Rendering {len(selected_brand_ids)} brands in defined area")
                brands = Brand.objects.filter(id__in=selected_brand_ids)
                print(f"🔍 Found {brands.count()} brands to render")

                # ADD DETAILED DEBUG
                for brand in brands:
                    print(f"🔍 Brand: {brand.name}, Image: {brand.brand_image}")
                    if brand.brand_image:
                        print(f"🔍 Brand image path: {brand.brand_image.path}")
                        print(f"🔍 Path exists: {os.path.exists(brand.brand_image.path)}")

                print(f"🔍 About to call render_brands_in_area with area: {template.brand_area_settings}")

                try:
                    # Ensure template_image is in RGBA mode before brand rendering
                    if template_image.mode != 'RGBA':
                        template_image = template_image.convert('RGBA')

                    self.render_brands_in_area(template_image, brands, template.brand_area_settings)
                    print("🔍 Successfully finished calling render_brands_in_area")
                except Exception as e:
                    print(f"🔍 ERROR in render_brands_in_area: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print("🔍 CONDITION FAILED - Brand rendering skipped")
                if not selected_brand_ids:
                    logger.info("No brands selected")
                    print("🔍 - No selected_brand_ids")
                elif not template.brand_area_settings:
                    logger.info("No brand area settings defined for template")
                    print("🔍 - No brand_area_settings")
                elif not template.brand_area_settings.get('enabled', False):
                    logger.info("Brand area not enabled for template")
                    print("🔍 - Brand area not enabled")

            # Save the final image AFTER all operations including brand rendering
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, f"temp_image_{uuid.uuid4().hex}.png")

            # Convert to RGB before saving to ensure compatibility
            # Save with transparency preserved - DON'T convert to RGB
            if template_image.mode == 'RGBA':
                template_image.save(output_path, 'PNG',
                                quality=100,
                                optimize=False,
                                compress_level=0,
                                dpi=(300, 300))
            else:
                template_image.save(output_path, 'PNG',
                                quality=100,
                                optimize=False,
                                compress_level=0,
                                dpi=(300, 300))

            return output_path
        #

        except Exception as e:
            logger.error(f"Image generation error: {e}")
            raise
        finally:
            # Comprehensive cleanup
            if template_image is not None:
                try:
                    template_image.close()
                    del template_image
                except:
                    pass
            if draw is not None:
                try:
                    del draw
                except:
                    pass
            # Clean up any brand images
            if hasattr(self, '_current_brand_images'):
                for brand_img in self._current_brand_images:
                    try:
                        brand_img.close()
                        del brand_img
                    except:
                        pass
                self._current_brand_images.clear()
                delattr(self, '_current_brand_images')

            # Force garbage collection
            import gc
            gc.collect()



    def render_brands_in_area(self, template_image, brands, area_settings):
        """Render brands using predefined slots with smart centering"""
        print(f"🔍 INSIDE render_brands_in_area with {brands.count()} brands")
        print(f"🔍 Area settings: {area_settings}")

        if not brands or not area_settings.get('enabled'):
            print("🔍 Exiting early - no brands or area not enabled")
            return

        area_x = area_settings.get('x', 50)
        area_y = area_settings.get('y', 400)
        area_width = area_settings.get('width', 700)
        area_height = area_settings.get('height', 150)
        slots = area_settings.get('slots', [])

        if not slots:
            print("🔍 No slots defined, exiting")
            return

        print(f"🔍 Area position: ({area_x}, {area_y})")
        print(f"🔍 Area size: {area_width}x{area_height}")
        print(f"🔍 Available slots: {len(slots)}")

        # Get only the slots we need (based on number of brands)
        brands_list = list(brands)
        needed_slots = slots[:len(brands_list)]

        print(f"🔍 Using {len(needed_slots)} slots for {len(brands_list)} brands")

        # Calculate bounding box of needed slots for centering
        if len(needed_slots) < len(slots):
            min_x = min(slot['x'] for slot in needed_slots)
            max_x = max(slot['x'] + slot['width'] for slot in needed_slots)
            min_y = min(slot['y'] for slot in needed_slots)
            max_y = max(slot['y'] + slot['height'] for slot in needed_slots)

            used_width = max_x - min_x
            used_height = max_y - min_y

            # Calculate centering offset
            center_offset_x = (area_width - used_width) // 2 - min_x
            center_offset_y = (area_height - used_height) // 2 - min_y

            print(f"🔍 Centering offset: ({center_offset_x}, {center_offset_y})")
        else:
            # Using all slots, no centering needed
            center_offset_x = 0
            center_offset_y = 0

        # Render brands in slots
        for i, (brand, slot) in enumerate(zip(brands_list, needed_slots)):
            print(f"🔍 Processing brand {i+1}: {brand.name}")

            if brand.brand_image and brand.brand_image.path and os.path.exists(brand.brand_image.path):
                try:
                    # Calculate final position with centering offset
                    final_x = area_x + slot['x'] + center_offset_x
                    final_y = area_y + slot['y'] + center_offset_y
                    slot_width = slot['width']
                    slot_height = slot['height']

                    print(f"🔍 Slot {i+1} position: ({final_x}, {final_y})")
                    print(f"🔍 Slot {i+1} size: {slot_width}x{slot_height}")

                    # Load and resize brand image
                    # Load and resize brand image with memory management
                    # Load and resize brand image with memory management
                    brand_img = None
                    try:
                        if not os.path.exists(brand.brand_image.path):
                            logger.warning(f"Brand image file not found: {brand.brand_image.path}")
                            continue


                        brand_img = Image.open(brand.brand_image.path)
                        # DEBUG: Check brand image properties
                        print(f"🔍 Brand {brand.name}:")
                        print(f"  - Mode: {brand_img.mode}")
                        print(f"  - Size: {brand_img.size}")
                        print(f"  - Has transparency: {brand_img.mode in ('RGBA', 'LA') or 'transparency' in brand_img.info}")
                        if brand_img.mode == 'RGBA':
                            # Check if image actually uses transparency
                            alpha = brand_img.split()[-1]
                            alpha_range = alpha.getextrema()
                            print(f"  - Alpha range: {alpha_range}")
                            print(f"  - Has real transparency: {alpha_range[0] < 255}")
                        if hasattr(self, '_current_brand_images'):
                            self._current_brand_images.append(brand_img)


                        # Keep transparency for brand images - DON'T convert to RGB
                        # Force RGBA mode to ensure transparency support
                        if brand_img.mode != 'RGBA':
                            if brand_img.mode == 'P' and 'transparency' in brand_img.info:
                                # Handle palette mode with transparency
                                brand_img = brand_img.convert('RGBA')
                            elif brand_img.mode in ('L', 'LA'):
                                # Handle grayscale with alpha
                                brand_img = brand_img.convert('RGBA')
                            else:
                                # Convert other modes to RGBA
                                brand_img = brand_img.convert('RGBA')

                        print(f"🔍 After conversion - Mode: {brand_img.mode}")
                        # Don't force RGB conversion - preserve transparency

                        # Resize brand image
                        # brand_img = brand_img.resize((slot_width, slot_height), Image.Resampling.LANCZOS)

                        # Resize brand image with better quality
                        original_ratio = brand_img.width / brand_img.height
                        slot_ratio = slot_width / slot_height

                        if original_ratio > slot_ratio:
                            # Image is wider - fit by width
                            new_width = slot_width
                            new_height = int(slot_width / original_ratio)
                        else:
                            # Image is taller - fit by height
                            new_height = slot_height
                            new_width = int(slot_height * original_ratio)

                        # Resize with high quality
                        brand_img = brand_img.resize((new_width, new_height), Image.Resampling.LANCZOS)

                        # Center the image in the slot if needed
                        # Center the image in the slot with transparent background
                        if new_width != slot_width or new_height != slot_height:
                            centered_img = Image.new('RGBA', (slot_width, slot_height), (0, 0, 0, 0))  # Fully transparent
                            paste_x = (slot_width - new_width) // 2
                            paste_y = (slot_height - new_height) // 2
                            if brand_img.mode == 'RGBA':
                                centered_img.paste(brand_img, (paste_x, paste_y), brand_img)  # Use alpha mask
                            else:
                                centered_img.paste(brand_img, (paste_x, paste_y))
                            brand_img = centered_img

                        # Paste brand image
                        # Paste brand image with proper alpha handling
                        if template_image.mode != 'RGBA':
                            template_image = template_image.convert('RGBA')

                        if brand_img.mode == 'RGBA':
                            template_image.paste(brand_img, (final_x, final_y), brand_img)  # Use alpha mask
                        else:
                            template_image.paste(brand_img, (final_x, final_y))

                    except (IOError, OSError) as e:
                        logger.warning(f"Failed to process brand image {brand.id}: {e}")
                        continue
                    finally:
                        if brand_img is not None:
                            try:
                                brand_img.close()
                            except:
                                pass

                    logger.info(f"Rendered brand {brand.name} in slot {i+1} at ({final_x}, {final_y})")
                    print(f"🔍 Successfully pasted brand {brand.name} at ({final_x}, {final_y})")

                except Exception as e:
                    print(f"🔍 Failed to render brand {brand.name}: {e}")
                    logger.error(f"Failed to render brand {brand.name}: {e}")

class ImageContentListView(APIView):
    """List generated image contents with pagination"""
    def get(self, request):
        doctor_id = request.GET.get('doctor_id')
        template_id = request.GET.get('template_id')
        contents = ImageContent.objects.all()
        if doctor_id:
            try:
                doctor = DoctorVideo.objects.get(id=doctor_id)
                contents = contents.filter(doctor=doctor)
            except DoctorVideo.DoesNotExist:
                return Response({"error": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)
        if template_id:
            contents = contents.filter(template_id=template_id)
        paginator = Pagination_class()
        page = paginator.paginate_queryset(contents, request)
        serializer = ImageContentSerializer(page, many=True, context={'request': request})
        return paginator.get_paginated_response(serializer.data)

class DoctorSearchView(APIView):
    """Search doctor by mobile number - Employee specific"""
    def get(self, request):
        mobile = request.GET.get('mobile')
        employee_id = request.GET.get('employee_id')
        
        if not mobile:
            return Response({"error": "mobile parameter required."}, status=status.HTTP_400_BAD_REQUEST)
            
        if not employee_id:
            return Response({"error": "employee_id parameter required."}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            current_employee = Employee.objects.get(employee_id=employee_id)
        except Employee.DoesNotExist:
            return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)
        # First check if current employee already has this doctor
        existing_doctor = DoctorVideo.objects.filter(
            mobile_number=mobile,
            employee=current_employee
        ).first()
        
        if existing_doctor:
            return Response({
                "found": True,
                "own_doctor": True,
                "doctor": {
                    "id": existing_doctor.id, "name": existing_doctor.name, "clinic": existing_doctor.clinic,
                    "city": existing_doctor.city, "mobile": existing_doctor.mobile_number,
                    "specialization": existing_doctor.specialization, "state": existing_doctor.state,
                }
            })
        
        # Check if any other employee has this doctor (for auto-population only)
        doctor_video = DoctorVideo.objects.filter(mobile_number=mobile).first()
        if doctor_video:
            return Response({
                "found": True,
                "own_doctor": False,
                "doctor": {
                    "name": doctor_video.name, "clinic": doctor_video.clinic,
                    "city": doctor_video.city, "mobile": doctor_video.mobile_number,
                    "specialization": doctor_video.specialization, "state": doctor_video.state,
                    "readonly_fields": ["name", "mobile_number"]  # These fields will be readonly
                }
            })
        
        return Response({"found": False})
    
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
# #@ratelimit(key='ip', rate='10/m', method='POST', block=True)
def AddEmployeeTemplates(request, template_type='video'):
    """Handle both video and image template creation"""
    try:
        data = request.data.copy()
        data['template_type'] = template_type

        if template_type == 'image':
            text_positions = data.get('text_positions', '{}')
            image_settings = data.get('imageSettings', '{}')
            if isinstance(text_positions, str):
                text_positions = json.loads(text_positions)
            if isinstance(image_settings, str):
                image_settings = json.loads(image_settings)
            combined_positions = text_positions.copy() if text_positions else {}
            combined_positions['imageSettings'] = image_settings
            data['text_positions'] = json.dumps(combined_positions)

        serializer = ImageTemplateSerializer(data=data, context={'request': request}) if template_type == 'image' \
            else VideoTemplatesSerializer(data=data)

        if serializer.is_valid():
            template = serializer.save()
            return Response({
                'id': template.id,
                'message': f'{template_type.title()} template created successfully',
                'status': template.status
            }, status=status.HTTP_201_CREATED)
        else:
            return Response({'error': 'Validation failed', 'details': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        return Response({'error': 'Template creation failed', 'details': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def getFilteredVideoTemplates(request):
    """Get filtered templates by status & type"""
    status_param = request.GET.get('status', 'true')
    template_type = request.GET.get('template_type', 'video')
    try:
        status_bool = bool(strtobool(status_param))
        templates = VideoTemplates.objects.filter(status=status_bool, template_type=template_type).order_by('-created_at')
        serializer = ImageTemplateSerializer(templates, many=True, context={'request': request}) if template_type == 'image' \
            else VideoTemplatesSerializer(templates, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ------------------------------------------------------------------------------
# Doctor Update/Delete (with permission checks)
# ------------------------------------------------------------------------------

class DoctorUpdateDeleteView(APIView):
    # permission_classes = []  # Explicitly override global auth
    permission_classes = [AllowAny]
    authentication_classes = []

    def get_doctor(self, doctor_id, employee_id):
        try:
            doctor = DoctorVideo.objects.get(id=doctor_id)
            try:
                current_employee = Employee.objects.get(employee_id=employee_id)
            except Employee.DoesNotExist:
                raise PermissionError("Employee not found")
            if current_employee.user_type != 'Admin' and doctor.employee != current_employee:
                raise PermissionError("You can only edit your own doctors")
            return doctor
        except DoctorVideo.DoesNotExist:
            raise Http404("Doctor not found")

    def patch(self, request, doctor_id):
        try:
            employee_id = request.data.get('employee_id')
            if not employee_id:
                return Response({'error': 'employee_id is required'}, status=status.HTTP_400_BAD_REQUEST)
            doctor = self.get_doctor(doctor_id, employee_id)
            serializer = DoctorVideoSerializer(doctor, data=request.data, partial=True)
            if serializer.is_valid():
                updated_doctor = serializer.save()
                return Response({'status': 'success', 'message': 'Doctor updated successfully', 'data': DoctorVideoSerializer(updated_doctor).data})
            else:
                return Response({'status': 'error', 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        except PermissionError as e:
            return Response({'error': str(e)}, status=status.HTTP_403_FORBIDDEN)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def delete(self, request, doctor_id):
        try:
            employee_id = request.query_params.get('employee_id')
            if not employee_id:
                return Response({'error': 'employee_id is required'}, status=status.HTTP_400_BAD_REQUEST)

            doctor = self.get_doctor(doctor_id, employee_id)
            doctor_name = doctor.name

            # DoctorOutputVideo model removed - only delete images
            deleted_videos = 0  # No video model to count
            deleted_images = ImageContent.objects.filter(doctor=doctor).count()

            # DoctorOutputVideo.objects.filter(doctor=doctor).delete()  # Model removed
            ImageContent.objects.filter(doctor=doctor).delete()

            # Delete the doctor
            doctor.delete()

            return Response({
                'status': 'success',
                'message': f'Doctor {doctor_name} deleted successfully. Removed {deleted_videos} videos and {deleted_images} images.',
                'deleted_content': {
                    'videos': deleted_videos,
                    'images': deleted_images
                }
            })
        except PermissionError as e:
            return Response({'error': str(e)}, status=status.HTTP_403_FORBIDDEN)
        except Exception as e:
            logger.error(f"Error deleting doctor {doctor_id}: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ------------------------------------------------------------------------------
# Delete specific content
# ------------------------------------------------------------------------------

class DeleteContentView(APIView):
    permission_classes = []
    authentication_classes = []

    def delete(self, request, content_type, content_id):
        try:
            if content_type == 'video':
                # DoctorOutputVideo model removed - video deletion not supported
                return Response({'error': 'Video content deletion not supported - model removed'}, status=status.HTTP_400_BAD_REQUEST)
            elif content_type == 'image':
                content = ImageContent.objects.get(id=content_id)
                doctor = content.doctor
            else:
                return Response({'error': 'Invalid content type'}, status=status.HTTP_400_BAD_REQUEST)

            employee_id = request.query_params.get('employee_id')
            if not employee_id:
                return Response({'error': 'employee_id is required'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                current_employee = Employee.objects.get(employee_id=employee_id)
            except Employee.DoesNotExist:
                return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
            if current_employee.user_type != 'Admin' and doctor.employee != current_employee:
                return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

            content.delete()
            return Response({'status': 'success', 'message': f'{content_type.title()} deleted successfully'})
        except ImageContent.DoesNotExist:
                    return Response({'error': 'Content not found'}, status=status.HTTP_404_NOT_FOUND)
        # DoctorOutputVideo.DoesNotExist removed - model doesn't exist
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class BrandListAPIView(generics.ListAPIView):
    queryset = Brand.objects.all().order_by('category', 'name')
    serializer_class = BrandSerializer
    permission_classes = [AllowAny]  # Add this line

    def get(self, request, *args, **kwargs):
        brands = self.get_queryset()

        # Group brands by category
        categories = {}
        for brand in brands:
            category = brand.category
            if category not in categories:
                categories[category] = {
                    'category_key': category,
                    'category_name': brand.get_category_display(),
                    'brands': []
                }
            categories[category]['brands'].append({
                'id': brand.id,
                'name': brand.name,
                'brand_image': request.build_absolute_uri(brand.brand_image.url) if brand.brand_image else None
            })

        return Response({
            'categories': list(categories.values()),
            'total_brands': brands.count()
        })

class ImageTemplateUsageView(APIView):
    def get(self, request):
        # Enhanced query with per-doctor usage counts
        template_data = ImageContent.objects.select_related('template', 'doctor').values(
            'template__id', 'template__name'
        ).annotate(
            usage_count=Count('id')
        ).filter(
            template__id__isnull=False
        ).order_by('-usage_count')

        data = []
        for item in template_data:
            # Get detailed doctor usage for this template
            doctor_usage = ImageContent.objects.filter(
                template_id=item["template__id"]
            ).values('doctor__name').annotate(
                count=Count('id')
            ).order_by('-count')

            doctor_details = [
                {
                    "name": usage["doctor__name"],
                    "usage_count": usage["count"]
                }
                for usage in doctor_usage if usage["doctor__name"]
            ]

            data.append({
                "template_id": item["template__id"],
                "template_name": item["template__name"],
                "usage_count": item["usage_count"],
                "doctor_names": [doc["name"] for doc in doctor_details],
                "doctor_details": doctor_details  # NEW: detailed usage per doctor
            })

        return Response(data, status=status.HTTP_200_OK)


from celery.result import AsyncResult

class TaskStatusView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, task_id):
        """Check image generation status"""
        result = AsyncResult(task_id)

        if result.ready():
            if result.successful():
                task_result = result.get()
                # Get the actual image content
                if 'image_id' in task_result:
                    try:
                        image_content = ImageContent.objects.get(id=task_result['image_id'])
                        serializer = ImageContentSerializer(image_content, context={'request': request})
                        return Response({
                            "status": "completed",
                            "result": serializer.data
                        })
                    except ImageContent.DoesNotExist:
                        pass

                return Response({"status": "completed", "result": task_result})
            else:
                return Response({"status": "failed", "error": str(result.info)})
        else:
            return Response({"status": "processing"})



from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def test_cors(request):
    return JsonResponse({"status": "CORS working", "method": request.method})

# class DoctorUsageHistoryView(APIView):
#     permission_classes = [AllowAny]

#     def get(self, request):
#         doctor_id = request.GET.get('doctor_id')
#         employee_id = request.GET.get('employee_id')

#         if doctor_id:
#             # Get usage history for specific doctor
#             history = DoctorUsageHistory.objects.filter(
#                 doctor_id=doctor_id,
#                 content_type='image'
#             ).select_related('employee', 'template')

#             return Response([{
#                 'employee_name': f"{h.employee.first_name} {h.employee.last_name}",
#                 'employee_id': h.employee.employee_id,
#                 'template_name': h.template.name,
#                 'generated_at': h.generated_at,
#                 'is_current_employee': h.employee.employee_id == employee_id
#             } for h in history])

#         return Response([])
    
# class SharedDoctorsView(APIView):
#     permission_classes = [AllowAny]

#     def get(self, request):
#         employee_id = request.GET.get('employee_id')

#         if not employee_id:
#             return Response({"error": "employee_id required"}, status=400)

#         try:
#             employee = Employee.objects.get(employee_id=employee_id)
#         except Employee.DoesNotExist:
#             return Response({"error": "Employee not found"}, status=404)

#         # Get doctors used by this employee but created by others
#         used_doctors = DoctorUsageHistory.objects.filter(
#             employee=employee,
#             content_type='image'
#         ).exclude(
#             doctor__employee=employee
#         ).select_related('doctor', 'doctor__employee').values('doctor').distinct()

#         # Get unique doctor IDs
#         doctor_ids = [usage['doctor'] for usage in used_doctors]
#         unique_doctors = DoctorVideo.objects.filter(id__in=doctor_ids)

#         doctors_data = []
#         for doctor in unique_doctors:
#             doctors_data.append({
#                 'id': doctor.id,
#                 'name': doctor.name,
#                 'mobile_number': doctor.mobile_number,
#                 'original_employee': f"{doctor.employee.first_name} {doctor.employee.last_name}",
#                 'last_used': DoctorUsageHistory.objects.filter(
#                     doctor=doctor,
#                     employee=employee
#                 ).latest('generated_at').generated_at,
#                 'usage_count': DoctorUsageHistory.objects.filter(
#                     doctor=doctor,
#                     employee=employee
#                 ).count()
#             })

#         return Response(doctors_data)